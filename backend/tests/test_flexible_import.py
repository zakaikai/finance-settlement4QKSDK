"""Tests for flexible import pipeline: parse Excel, resolve game names, import data."""
import os
import tempfile
import pytest
from datetime import date
from decimal import Decimal
from openpyxl import Workbook
from sqlalchemy import select

from backend import models
from backend.services.flexible_import import (
    parse_flexible_excel,
    resolve_flexible_game_names,
    import_flexible_data,
    _to_decimal,
)


# ── Helper: create temp Excel file ──

def _make_xlsx(headers: list, data_rows: list[list], header_row: int = 1) -> str:
    """Create a temporary .xlsx file with given headers and data rows. Returns path."""
    wb = Workbook()
    ws = wb.active
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    for row_idx, row in enumerate(data_rows, start=header_row + 1):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    wb.close()
    return tmp.name


# ── Helper: seed minimal channel hierarchy + game ──

async def _seed_channel(db, channel_id=1, channel_name=None):
    """Seed only channel hierarchy (no game)."""
    if channel_name is None:
        channel_name = f"测试渠道{channel_id}"
    db.add(models.ChannelCategory(channel_id=channel_id, channel_name=channel_name))
    db.add(models.BackendChannel(backend_channel_id=channel_id, backend_channel_name=f"测试后台{channel_id}", channel_id=channel_id))
    db.add(models.SubChannel(sub_channel_id=channel_id, sub_channel_name=f"测试子渠道{channel_id}", backend_channel_id=channel_id))
    await db.commit()


async def _seed_game(db, game_id="G001", game_name="测试游戏", discount_rate=None):
    """Seed a single game."""
    if discount_rate is None:
        discount_rate = Decimal("0.8")
    db.add(models.Game(game_id=game_id, game_name=game_name, discount_rate=discount_rate))
    await db.commit()


async def _seed_channel_game(db, channel_id=1, game_id="G001", game_name="测试游戏",
                           channel_name=None):
    """Seed ChannelCategory + BackendChannel + SubChannel + Game."""
    await _seed_channel(db, channel_id, channel_name)
    await _seed_game(db, game_id, game_name)
    await db.commit()


# ═══════════════════════════════════════════════════════════════
# A. parse_flexible_excel
# ═══════════════════════════════════════════════════════════════

class TestParseFlexibleExcel:
    """Tests for parse_flexible_excel — Excel parsing with column mapping."""

    @pytest.mark.asyncio
    async def test_parse_basic_rows(self):
        """Parse two data rows with game_name and money fields mapped."""
        path = _make_xlsx(
            ["游戏", "流水金额", "代金券"],
            [["王者荣耀", 10000, 500], ["和平精英", 20000, 300]],
        )
        mapping = {"0": "game_name", "1": "raw_revenue", "2": "vouchers"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["total_rows"] == 2
            assert len(result["errors"]) == 0
            assert result["rows"][0]["game_name"] == "王者荣耀"
            assert result["rows"][0]["raw_revenue"] == Decimal("10000")
            assert result["rows"][0]["vouchers"] == Decimal("500")
            assert result["rows"][1]["game_name"] == "和平精英"
            assert result["rows"][1]["raw_revenue"] == Decimal("20000")
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_money_with_chinese_symbols(self):
        """Money fields strip ¥, ￥, 元, and commas."""
        path = _make_xlsx(
            ["流水"],
            [["¥1,234.56"], ["￥7,890"], ["100元"], ["5,000.00"]],
        )
        mapping = {"0": "raw_revenue"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["total_rows"] == 4
            assert result["rows"][0]["raw_revenue"] == Decimal("1234.56")
            assert result["rows"][1]["raw_revenue"] == Decimal("7890")
            assert result["rows"][2]["raw_revenue"] == Decimal("100")
            assert result["rows"][3]["raw_revenue"] == Decimal("5000.00")
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_invalid_money_value(self):
        """Invalid money values produce row errors and None in record."""
        path = _make_xlsx(
            ["流水", "名称"],
            [["abc", "游戏A"], ["100", "游戏B"]],
        )
        mapping = {"0": "raw_revenue", "1": "game_name"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert len(result["errors"]) == 1
            assert "无法转为数字" in result["errors"][0]["errors"][0]
            # Row with error is excluded from rows
            assert result["total_rows"] == 1
            assert result["rows"][0]["raw_revenue"] == Decimal("100")
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_skips_empty_rows(self):
        """Entirely blank rows (all None) are skipped."""
        path = _make_xlsx(
            ["流水"],
            [[100], [None], [200]],
        )
        mapping = {"0": "raw_revenue"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["total_rows"] == 2
            assert result["rows"][0]["raw_revenue"] == Decimal("100")
            assert result["rows"][1]["raw_revenue"] == Decimal("200")
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_custom_header_row(self):
        """Header on row 3, header_row=3 — data starts at row 4."""
        # Manual construction: rows 1-2 are garbage, row 3 is header, row 4 is data
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ignored row")
        ws.cell(row=2, column=1, value="also ignored")
        ws.cell(row=3, column=1, value="流水")  # header
        ws.cell(row=4, column=1, value=999)      # data
        path = os.path.join(tempfile.gettempdir(), f"test_hdr_{id(self)}.xlsx")
        wb.save(path)
        wb.close()
        try:
            result = await parse_flexible_excel(path, {"0": "raw_revenue"}, header_row=3)
            assert result["total_rows"] == 1
            assert result["rows"][0]["raw_revenue"] == Decimal("999")
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass

    @pytest.mark.asyncio
    async def test_parse_empty_cell_produces_none(self):
        """Empty/blank cell in Excel maps to None in record."""
        path = _make_xlsx(
            ["名称", "流水"],
            [["游戏A", None], [None, 500]],
        )
        mapping = {"0": "game_name", "1": "raw_revenue"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["total_rows"] == 2
            assert result["rows"][0]["game_name"] == "游戏A"
            assert result["rows"][0]["raw_revenue"] is None
            assert result["rows"][1]["game_name"] is None
            assert result["rows"][1]["raw_revenue"] == Decimal("500")
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_ignore_column(self):
        """Columns mapped to 'ignore' are excluded from records."""
        path = _make_xlsx(
            ["忽略列", "流水"],
            [["whatever", 100]],
        )
        mapping = {"0": "ignore", "1": "raw_revenue"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["total_rows"] == 1
            assert "ignore" not in result["rows"][0]
            assert result["rows"][0]["raw_revenue"] == Decimal("100")
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_non_money_string_field(self):
        """Non-money fields are stored as stripped strings."""
        path = _make_xlsx(
            ["游戏名", "月份"],
            [["  王者荣耀  ", "2026-04"]],
        )
        mapping = {"0": "game_name", "1": "month"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["rows"][0]["game_name"] == "王者荣耀"
            assert result["rows"][0]["month"] == "2026-04"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_row_with_shorter_length(self):
        """Row shorter than mapped column index gets None for out-of-bounds col."""
        path = _make_xlsx(
            ["流水"],
            [[100]],
        )
        # Map column index 5 which doesn't exist in this row
        mapping = {"0": "raw_revenue", "5": "vouchers"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert result["rows"][0]["raw_revenue"] == Decimal("100")
            assert result["rows"][0]["vouchers"] is None
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_decimal_money(self):
        """Float values in Excel are parsed as Decimal correctly."""
        path = _make_xlsx(
            ["流水"],
            [[1234.56]],
        )
        mapping = {"0": "raw_revenue"}
        try:
            result = await parse_flexible_excel(path, mapping)
            assert isinstance(result["rows"][0]["raw_revenue"], Decimal)
            assert result["rows"][0]["raw_revenue"] == Decimal("1234.56")
        finally:
            os.unlink(path)


# ═══════════════════════════════════════════════════════════════
# B. resolve_flexible_game_names
# ═══════════════════════════════════════════════════════════════

class TestResolveFlexibleGameNames:
    """Tests for resolve_flexible_game_names — fuzzy game name matching in flex import."""

    @pytest.mark.asyncio
    async def test_exact_match_high_confidence(self, db_session):
        """Exact game name match returns high confidence."""
        db_session.add(models.Game(game_id="G001", game_name="王者荣耀", discount_rate=Decimal("0.8")))
        await db_session.commit()

        rows = [{"game_name": "王者荣耀"}]
        results, errors = await resolve_flexible_game_names(db_session, rows, "game_name")

        assert len(errors) == 0
        assert rows[0]["game_id"] == "G001"
        assert rows[0]["match_status"] == "high"
        assert rows[0]["matched_game_name"] == "王者荣耀"

    @pytest.mark.asyncio
    async def test_fuzzy_match_medium_or_high(self, db_session):
        """Slightly misspelled game name still matches."""
        db_session.add(models.Game(game_id="G001", game_name="王者荣耀", discount_rate=Decimal("0.8")))
        await db_session.commit()

        rows = [{"game_name": "王者荣罐"}]  # one char different
        results, errors = await resolve_flexible_game_names(db_session, rows, "game_name")

        assert rows[0]["game_id"] == "G001"
        assert rows[0]["match_status"] in ("high", "medium")

    @pytest.mark.asyncio
    async def test_empty_name_produces_error(self, db_session):
        """Empty game_name produces an error entry."""
        rows = [{"game_name": ""}]
        results, errors = await resolve_flexible_game_names(db_session, rows, "game_name")

        assert len(errors) == 1
        assert "为空" in errors[0]["error"]

    @pytest.mark.asyncio
    async def test_multiple_rows_resolved(self, db_session):
        """Multiple rows with different game names all resolve."""
        db_session.add(models.Game(game_id="G001", game_name="王者荣耀", discount_rate=Decimal("0.8")))
        db_session.add(models.Game(game_id="G002", game_name="和平精英", discount_rate=Decimal("0.7")))
        await db_session.commit()

        rows = [{"game_name": "王者荣耀"}, {"game_name": "和平精英"}]
        results, errors = await resolve_flexible_game_names(db_session, rows, "game_name")

        assert len(results) == 2
        assert rows[0]["game_id"] == "G001"
        assert rows[1]["game_id"] == "G002"

    @pytest.mark.asyncio
    async def test_missing_game_name_key(self, db_session):
        """Row without the game_name_col key treated as empty."""
        db_session.add(models.Game(game_id="G001", game_name="测试", discount_rate=Decimal("0.8")))
        await db_session.commit()

        rows = [{"other_field": "value"}]
        results, errors = await resolve_flexible_game_names(db_session, rows, "game_name")

        assert len(errors) == 1
        assert "为空" in errors[0]["error"]

    @pytest.mark.asyncio
    async def test_game_name_none_value(self, db_session):
        """None game_name value treated as empty string."""
        rows = [{"game_name": None}]
        results, errors = await resolve_flexible_game_names(db_session, rows, "game_name")

        assert len(errors) == 1
        assert "为空" in errors[0]["error"]


# ═══════════════════════════════════════════════════════════════
# C. import_flexible_data
# ═══════════════════════════════════════════════════════════════

class TestImportFlexibleData:
    """Tests for import_flexible_data — the DB write path."""

    @pytest.mark.asyncio
    async def test_import_updates_deduction(self, db_session):
        """Existing deduction row is updated with imported values."""
        await _seed_channel_game(db_session)
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-04",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        await db_session.commit()

        rows = [{
            "game_id": "G001", "game_name": "测试游戏",
            "vouchers": 100, "test": 0, "welfare": 0, "bad_debt": 50,
        }]
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                            column_mapping=mapping)
        await db_session.commit()

        assert result["imported_deductions"] == 1
        ded = (await db_session.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == 1,
                models.Deduction.game_id == "G001",
                models.Deduction.month == "2026-04",
            )
        )).scalar_one()
        assert float(ded.vouchers) == 100.0
        assert float(ded.bad_debt) == 50.0

    @pytest.mark.asyncio
    async def test_import_skips_without_existing_deduction(self, db_session):
        """Row without existing Deduction: deduction write is skipped, but lock/config still checked."""
        await _seed_channel_game(db_session)

        rows = [{"game_id": "G001", "vouchers": 100}]
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                            column_mapping=mapping)
        await db_session.commit()

        assert result["imported_deductions"] == 0

    @pytest.mark.asyncio
    async def test_import_updates_existing_deduction(self, db_session):
        """Existing Deduction for same (channel, game, month) is updated in-place."""
        await _seed_channel_game(db_session)
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-04",
            vouchers=Decimal("50"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        await db_session.commit()

        rows = [{
            "game_id": "G001", "game_name": "测试游戏",
            "vouchers": 200, "test": 10, "welfare": 20, "bad_debt": 30,
        }]
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                            column_mapping=mapping)
        await db_session.commit()

        ded = (await db_session.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == 1,
                models.Deduction.game_id == "G001",
                models.Deduction.month == "2026-04",
            )
        )).scalar_one()
        assert float(ded.vouchers) == 200.0
        assert float(ded.test) == 10.0

    @pytest.mark.asyncio
    async def test_import_blocked_by_existing_lock(self, db_session):
        """Frozen guard: import rejected when ChannelLock exists for (channel, game, month)."""
        await _seed_channel_game(db_session)
        db_session.add(models.ChannelLock(
            channel_id=1, game_id="G001", month="2026-05",
            locked_real_revenue=Decimal("3000"),
            created_at="old", updated_at="old",
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "raw_revenue": 8000}]
        mapping = {}
        with pytest.raises(ValueError, match="已被锁定"):
            await import_flexible_data(db_session, rows, channel_id=1, month="2026-05",
                                       column_mapping=mapping)

    @pytest.mark.asyncio
    async def test_import_blocked_by_existing_lock_settlement(self, db_session):
        """Frozen guard blocks import when ChannelLock with settlement_amount exists."""
        await _seed_channel_game(db_session)
        db_session.add(models.ChannelLock(
            channel_id=1, game_id="G001", month="2026-06",
            locked_settlement_amount=Decimal("1000"),
            created_at="old", updated_at="old",
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "settlement_amount": 4200}]
        mapping = {}
        with pytest.raises(ValueError, match="已被锁定"):
            await import_flexible_data(db_session, rows, channel_id=1, month="2026-06",
                                       column_mapping=mapping)

    @pytest.mark.asyncio
    async def test_import_blocked_when_lock_exists(self, db_session):
        """Frozen guard blocks import when ChannelLock exists, regardless of fields."""
        await _seed_channel_game(db_session)
        db_session.add(models.ChannelLock(
            channel_id=1, game_id="G001", month="2026-07",
            locked_real_revenue=Decimal("3000"),
            locked_settlement_amount=Decimal("2000"),
            created_at="old", updated_at="old",
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "raw_revenue": 9999}]
        mapping = {}
        with pytest.raises(ValueError, match="已被锁定"):
            await import_flexible_data(db_session, rows, channel_id=1, month="2026-07",
                                       column_mapping=mapping)

    @pytest.mark.asyncio
    async def test_import_creates_split_config(self, db_session):
        """When split_rate/channel_fee_rate/tax_rate present, IncomeSplitConfig is created."""
        await _seed_channel_game(db_session)
        # Need pre-existing Deduction for the row to be accepted
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-04",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "split_rate": 0.5, "channel_fee_rate": 0.05, "tax_rate": 0.06}]
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                            column_mapping=mapping)
        await db_session.commit()

        assert result["imported_configs"] == 1
        config = (await db_session.execute(
            select(models.IncomeSplitConfig).where(
                models.IncomeSplitConfig.channel_id == 1,
                models.IncomeSplitConfig.game_id == "G001",
            )
        )).scalar_one()
        assert float(config.split_rate) == 0.5
        assert float(config.channel_fee_rate) == 0.05

    @pytest.mark.asyncio
    async def test_import_duplicate_game_month_raises(self, db_session):
        """Two rows with same (game_id, month) raise ValueError."""
        await _seed_channel_game(db_session)

        rows = [
            {"game_id": "G001", "vouchers": 100},
            {"game_id": "G001", "vouchers": 200},  # same game + month
        ]
        mapping = {}
        with pytest.raises(ValueError, match="重复匹配的游戏条目"):
            await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                       column_mapping=mapping)

    @pytest.mark.asyncio
    async def test_import_selected_indices_with_global_month(self, db_session):
        """selected_indices correctly filters which rows are imported when global month used."""
        await _seed_channel_game(db_session)
        # Seed deductions for all three (game, month) pairs so updates succeed
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-04",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        # Create a second game for row 2
        await _seed_game(db_session, game_id="G002", game_name="游戏B")
        db_session.add(models.Deduction(
            channel_id=1, game_id="G002", month="2026-04",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        await db_session.commit()

        rows = [
            {"game_id": "G001", "vouchers": 111},
            {"game_id": "G001", "vouchers": 222},  # skipped by selected_indices
            {"game_id": "G002", "vouchers": 333},
        ]
        mapping = {}
        result = await import_flexible_data(
            db_session, rows, channel_id=1, month="2026-04",
            column_mapping=mapping,
            selected_indices={0, 2},
        )
        await db_session.commit()

        assert result["imported_deductions"] == 2
        # Row 0 updated G001
        ded1 = (await db_session.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == 1,
                models.Deduction.game_id == "G001",
                models.Deduction.month == "2026-04",
            )
        )).scalar_one()
        assert float(ded1.vouchers) == 111.0
        # Row 2 updated G002
        ded2 = (await db_session.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == 1,
                models.Deduction.game_id == "G002",
                models.Deduction.month == "2026-04",
            )
        )).scalar_one()
        assert float(ded2.vouchers) == 333.0

    @pytest.mark.asyncio
    async def test_import_real_revenue_blocked_by_lock(self, db_session):
        """real_revenue fallback for lock is blocked by frozen guard when lock exists."""
        await _seed_channel_game(db_session)
        db_session.add(models.ChannelLock(
            channel_id=1, game_id="G001", month="2026-08",
            locked_real_revenue=Decimal("1000"),
            created_at="old", updated_at="old",
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "real_revenue": 7500}]
        mapping = {}
        with pytest.raises(ValueError, match="已被锁定"):
            await import_flexible_data(db_session, rows, channel_id=1, month="2026-08",
                                       column_mapping=mapping)

    @pytest.mark.asyncio
    async def test_import_zero_value_blocked_by_lock(self, db_session):
        """Zero value import is blocked when ChannelLock exists."""
        await _seed_channel_game(db_session)
        db_session.add(models.ChannelLock(
            channel_id=1, game_id="G001", month="2026-09",
            locked_real_revenue=Decimal("500"),
            created_at="old", updated_at="old",
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "raw_revenue": 0}]
        mapping = {}
        with pytest.raises(ValueError, match="已被锁定"):
            await import_flexible_data(db_session, rows, channel_id=1, month="2026-09",
                                       column_mapping=mapping)

    @pytest.mark.asyncio
    async def test_import_no_lock_when_both_none(self, db_session):
        """When neither raw_revenue nor settlement_amount given, no lock created."""
        await _seed_channel_game(db_session)

        rows = [{"game_id": "G001", "vouchers": 100}]
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-10",
                                            column_mapping=mapping)
        await db_session.commit()

        lock = (await db_session.execute(
            select(models.ChannelLock).where(
                models.ChannelLock.channel_id == 1,
                models.ChannelLock.game_id == "G001",
                models.ChannelLock.month == "2026-10",
            )
        )).scalar_one_or_none()
        assert lock is None

    @pytest.mark.asyncio
    async def test_import_skips_row_without_game_id(self, db_session):
        """Row without game_id is silently skipped."""
        await _seed_channel_game(db_session)

        rows = [{"vouchers": 100}]  # no game_id
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                            column_mapping=mapping)
        await db_session.commit()

        assert result["imported_deductions"] == 0

    @pytest.mark.asyncio
    async def test_import_skips_row_without_month(self, db_session):
        """Row without any month source is silently skipped."""
        await _seed_channel_game(db_session)

        rows = [{"game_id": "G001", "vouchers": 100}]  # no month in record
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="",
                                            column_mapping=mapping)
        await db_session.commit()

        assert result["imported_deductions"] == 0

    @pytest.mark.asyncio
    async def test_import_returns_correct_counts(self, db_session):
        """Return dict has accurate counts for deductions and configs."""
        await _seed_channel_game(db_session)

        rows = [
            {"game_id": "G001", "vouchers": 100, "split_rate": 0.5},
            {"game_id": "G001", "vouchers": 200, "channel_fee_rate": 0.03},
        ]
        mapping = {}
        # Use different months to avoid duplicate guard, same game
        # Wait — same (game_id, month) with global month "2026-11" → duplicate!
        # Need to use row-level months
        rows = [
            {"game_id": "G001", "month": "2026-11", "vouchers": 100, "split_rate": 0.5},
            {"game_id": "G001", "month": "2026-12", "vouchers": 200, "channel_fee_rate": 0.03},
        ]
        # Seed existing deductions
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-11",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-12",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        await db_session.commit()

        result = await import_flexible_data(db_session, rows, channel_id=1, month="",
                                            column_mapping=mapping)
        await db_session.commit()

        assert result["imported_deductions"] == 2
        assert result["imported_configs"] == 2

    @pytest.mark.asyncio
    async def test_import_deduction_missing_fields_default_to_zero(self, db_session):
        """Missing deduction fields keep their existing values."""
        await _seed_channel_game(db_session)
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-11",
            vouchers=Decimal("50"), test=Decimal("10"),
            welfare=Decimal("20"), bad_debt=Decimal("30"),
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "vouchers": 500}]  # test/welfare/bad_debt not provided
        mapping = {}
        result = await import_flexible_data(db_session, rows, channel_id=1, month="2026-11",
                                            column_mapping=mapping)
        await db_session.commit()

        ded = (await db_session.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == 1,
                models.Deduction.game_id == "G001",
                models.Deduction.month == "2026-11",
            )
        )).scalar_one()
        assert float(ded.vouchers) == 500.0
        # Unmapped fields keep their pre-existing values
        assert float(ded.test) == 10.0
        assert float(ded.welfare) == 20.0
        assert float(ded.bad_debt) == 30.0

    @pytest.mark.asyncio
    async def test_import_different_channels_independent(self, db_session):
        """Same (game, month) for different channels updates separate deductions."""
        await _seed_channel_game(db_session, channel_id=1)
        await _seed_channel(db_session, channel_id=2)
        # Pre-seed deductions for both channels
        db_session.add(models.Deduction(
            channel_id=1, game_id="G001", month="2026-04",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        db_session.add(models.Deduction(
            channel_id=2, game_id="G001", month="2026-04",
            vouchers=Decimal("0"), test=Decimal("0"),
            welfare=Decimal("0"), bad_debt=Decimal("0"),
        ))
        await db_session.commit()

        rows = [{"game_id": "G001", "vouchers": 100}]
        mapping = {}
        result1 = await import_flexible_data(db_session, rows, channel_id=1, month="2026-04",
                                             column_mapping=mapping)
        result2 = await import_flexible_data(db_session, rows, channel_id=2, month="2026-04",
                                             column_mapping=mapping)
        await db_session.commit()

        assert result1["imported_deductions"] == 1
        assert result2["imported_deductions"] == 1
        # Two separate deduction rows
        count = (await db_session.execute(
            select(models.Deduction).where(
                models.Deduction.game_id == "G001",
                models.Deduction.month == "2026-04",
            )
        )).scalars().all()
        assert len(count) == 2


# ═══════════════════════════════════════════════════════════════
# D. _to_decimal helper
# ═══════════════════════════════════════════════════════════════

class TestToDecimal:
    """Tests for _to_decimal — the safe Decimal converter."""

    def test_valid_number(self):
        assert _to_decimal("123.45") == Decimal("123.45")

    def test_valid_int(self):
        assert _to_decimal(100) == Decimal("100")

    def test_none_returns_none(self):
        assert _to_decimal(None) is None

    def test_negative_returns_none(self):
        assert _to_decimal("-50") is None

    def test_invalid_string_returns_none(self):
        assert _to_decimal("abc") is None

    def test_zero_returns_zero(self):
        assert _to_decimal("0") == Decimal("0")
        assert _to_decimal(0) == Decimal("0")
