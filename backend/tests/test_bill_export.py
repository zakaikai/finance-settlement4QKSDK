"""Tests for bill export Excel generation."""
import pytest
from io import BytesIO
from types import SimpleNamespace
from decimal import Decimal
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from backend.main import app
from backend.database import get_db
from backend.services.bill_service import generate_channel_bill, generate_publisher_bill


def _make_party(**kwargs):
    defaults = dict(
        name="测试公司",
        address="北京市朝阳区测试路100号",
        phone="010-88888888",
        bank_name="中国银行北京测试支行",
        bank_account="6222021234567890",
        tax_id="91110000MA11111111",
    )
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _read_excel(buf: BytesIO):
    buf.seek(0)
    return load_workbook(buf)


@pytest.mark.asyncio
async def test_channel_bill_contains_party_info():
    """Channel bill Excel includes 甲方 and 乙方 details."""
    party_a = _make_party(name="我方科技有限公司", bank_account="6222021111111111")
    party_b = _make_party(name="华为渠道", bank_name="招商银行深圳分行")

    buf = await generate_channel_bill(party_a, "2026-03", party_b, [])
    wb = _read_excel(buf)
    ws = wb.active

    cells = [ws.cell(row=r, column=1).value for r in range(1, 30)]
    text = "\n".join(str(c or "") for c in cells)

    assert "我方科技有限公司" in text
    assert "6222021111111111" in text
    assert "华为渠道" in text
    assert "招商银行深圳分行" in text
    assert "收 入 结 算 对 账 单" in text
    assert "2026年3月" in text or "2026-03" in text


@pytest.mark.asyncio
async def test_publisher_bill_contains_party_info():
    """Publisher bill Excel includes 甲方 and 乙方 details."""
    party_a = _make_party(name="我方公司", tax_id="91110000MA99999999")
    party_b = _make_party(name="某研发工作室", bank_name="建设银行")

    buf = await generate_publisher_bill(party_a, "2026-04", party_b, [])
    wb = _read_excel(buf)
    ws = wb.active

    cells = [ws.cell(row=r, column=1).value for r in range(1, 30)]
    text = "\n".join(str(c or "") for c in cells)

    assert "我方公司" in text
    assert "91110000MA99999999" in text
    assert "某研发工作室" in text
    assert "建设银行" in text
    assert "付 款 结 算 对 账 单" in text
    assert "2026年4月" in text or "2026-04" in text


@pytest.mark.asyncio
async def test_channel_bill_empty_data_has_headers():
    """Channel bill with no data still produces valid Excel with header row."""
    party_a = _make_party()
    party_b = _make_party(name="渠道乙")

    buf = await generate_channel_bill(party_a, "2026-03", party_b, [])
    wb = _read_excel(buf)
    ws = wb.active

    header_texts = [ws.cell(row=r, column=c).value
                    for r in range(1, 40)
                    for c in range(1, 19)
                    if ws.cell(row=r, column=c).value]
    joined = "|".join(str(h or "") for h in header_texts)

    assert "序号" in joined
    assert "游戏编号" in joined
    assert "游戏名称" in joined
    assert "原始流水" in joined
    assert "结算金额" in joined
    assert "合  计" in joined


@pytest.mark.asyncio
async def test_channel_bill_summary_totals():
    """Summary row correctly sums numeric columns."""
    party_a = _make_party()
    party_b = _make_party(name="渠道")
    rows = [
        {"game_id": "G001", "game_name": "Game1", "project_code": "", "project_name": "",
         "raw_revenue": 1000, "real_revenue": 900, "vouchers": 50, "test": 20,
         "welfare": 10, "bad_debt": 5, "total_deductions": 85,
         "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
         "settlement_amount": 348.84},
        {"game_id": "G002", "game_name": "Game2", "project_code": "", "project_name": "",
         "raw_revenue": 2000, "real_revenue": 1800, "vouchers": 100, "test": 40,
         "welfare": 20, "bad_debt": 10, "total_deductions": 170,
         "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
         "settlement_amount": 697.68},
    ]

    buf = await generate_channel_bill(party_a, "2026-03", party_b, rows)
    wb = _read_excel(buf)
    ws = wb.active

    # Find the summary row by scanning for "合  计"
    summary_row = None
    for r in range(1, 50):
        if ws.cell(row=r, column=1).value == "合  计":
            summary_row = r
            break
    assert summary_row is not None

    # Column 7 = raw_revenue, Column 8 = real_revenue, Column 18 = settlement_amount
    assert ws.cell(row=summary_row, column=7).value == 3000  # 1000 + 2000
    assert ws.cell(row=summary_row, column=8).value == 2700  # 900 + 1800
    assert ws.cell(row=summary_row, column=18).value == 1046.52  # 348.84 + 697.68


@pytest.mark.asyncio
async def test_bill_api_with_filtered_rows(db_session):
    """POST /api/settlement/bill 传入过滤后的行数据, Excel 只包含这些行."""
    from backend import models

    # Seed PartyInfo
    party_a = models.PartyInfo(
        party_type="our_company", name="我方公司", address="addr",
        bank_name="银行A", bank_account="acctA", tax_id="taxA",
    )
    party_b = models.PartyInfo(
        party_type="channel", name="渠道乙", address="addr",
        bank_name="银行B", bank_account="acctB", tax_id="taxB",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    try:
        rows = [
            {"game_id": "G001", "game_name": "Game1", "project_code": "", "project_name": "",
             "raw_revenue": 1000, "real_revenue": 900, "vouchers": 50, "test": 20,
             "welfare": 10, "bad_debt": 5, "total_deductions": 85,
             "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
             "settlement_amount": 348.84},
            {"game_id": "G002", "game_name": "Game2", "project_code": "", "project_name": "",
             "raw_revenue": 2000, "real_revenue": 1800, "vouchers": 100, "test": 40,
             "welfare": 20, "bad_debt": 10, "total_deductions": 170,
             "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
             "settlement_amount": 697.68},
        ]
        resp = client.post("/api/settlement/bill", json={
            "mode": "income",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-03",
            "end_month": "2026-03",
            "rows": rows,
        })
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Party info and period appear — scan all cells
        all_vals = []
        for r in range(1, 50):
            for c in range(1, 18):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_vals.append(str(v))
        text = "|".join(all_vals)
        assert "我方公司" in text
        assert "2026年3月" in text or "2026-03" in text
        assert "G001" in text
        assert "G002" in text
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_bill_api_empty_rows_produces_valid_excel(db_session):
    """POST /api/settlement/bill 传入空 rows，仍然生成有效 Excel（含表头）。"""
    from backend import models
    from backend.database import get_db
    from backend.main import app
    from fastapi.testclient import TestClient

    party_a = models.PartyInfo(
        party_type="our_company", name="我方公司", address="addr",
        bank_name="银行A", bank_account="acctA", tax_id="taxA",
    )
    party_b = models.PartyInfo(
        party_type="channel", name="渠道乙", address="addr",
        bank_name="银行B", bank_account="acctB", tax_id="taxB",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override
    try:
        client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
        resp = client.post("/api/settlement/bill", json={
            "mode": "income",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-03",
            "end_month": "2026-03",
            "rows": [],  # empty
        })
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Should at least have 甲方/乙方 info and column headers
        all_vals = []
        for r in range(1, 40):
            for c in range(1, 18):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_vals.append(str(v))
        text = "|".join(all_vals)
        assert "我方公司" in text
        assert "渠道乙" in text
        assert "合  计" in text
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_bill_api_publisher_filtered_rows(db_session):
    """POST /api/settlement/bill publisher 模式 + 过滤行。"""
    from backend import models
    from backend.database import get_db
    from backend.main import app
    from fastapi.testclient import TestClient

    party_a = models.PartyInfo(
        party_type="our_company", name="我方公司", address="addr",
        bank_name="银行A", bank_account="acctA", tax_id="taxA",
    )
    party_b = models.PartyInfo(
        party_type="publisher", name="研发工作室", address="addr",
        bank_name="银行C", bank_account="acctC", tax_id="taxC",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override
    try:
        client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
        rows = [
            {"game_id": "G001", "game_name": "Game1", "project_code": "", "project_name": "",
             "raw_revenue": 1000, "real_revenue": 900, "split_rate": 0.5, "channel_fee_rate": 0.1,
             "tax_rate": 0.05, "settlement_amount": 348.84},
        ]
        resp = client.post("/api/settlement/bill", json={
            "mode": "payment",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-04",
            "end_month": "2026-04",
            "rows": rows,
        })
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        all_vals = []
        for r in range(1, 40):
            for c in range(1, 18):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_vals.append(str(v))
        text = "|".join(all_vals)
        assert "研发工作室" in text
        assert "2026年4月" in text or "2026-04" in text
        assert "G001" in text
        assert "付 款 结 算 对 账 单" in text
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_bill_api_filtered_rows_summary_totals(db_session):
    """Excel 合计行汇总只包含传入过滤行。"""
    from backend import models
    from backend.database import get_db
    from backend.main import app
    from fastapi.testclient import TestClient

    party_a = models.PartyInfo(
        party_type="our_company", name="我方公司", address="addr",
        bank_name="银行A", bank_account="acctA", tax_id="taxA",
    )
    party_b = models.PartyInfo(
        party_type="channel", name="渠道乙", address="addr",
        bank_name="银行B", bank_account="acctB", tax_id="taxB",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override
    try:
        client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
        rows = [
            {"game_id": "G001", "game_name": "Game1", "project_code": "", "project_name": "",
             "raw_revenue": 1000, "real_revenue": 900, "vouchers": 50, "test": 20,
             "welfare": 10, "bad_debt": 5, "total_deductions": 85,
             "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
             "settlement_amount": 348.84},
            {"game_id": "G002", "game_name": "Game2", "project_code": "", "project_name": "",
             "raw_revenue": 2000, "real_revenue": 1800, "vouchers": 100, "test": 40,
             "welfare": 20, "bad_debt": 10, "total_deductions": 170,
             "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
             "settlement_amount": 697.68},
        ]
        resp = client.post("/api/settlement/bill", json={
            "mode": "income",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-03",
            "end_month": "2026-03",
            "rows": rows,
        })
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active

        # Find summary row
        summary_row = None
        for r in range(1, 50):
            if ws.cell(row=r, column=1).value == "合  计":
                summary_row = r
                break
        assert summary_row is not None
        assert ws.cell(row=summary_row, column=7).value == 3000   # raw_revenue sum
        assert ws.cell(row=summary_row, column=8).value == 2700   # real_revenue sum
        assert ws.cell(row=summary_row, column=18).value == 1046.52  # settlement_amount sum
    finally:
        app.dependency_overrides.clear()
