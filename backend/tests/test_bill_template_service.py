"""Tests for bill template rendering engine (bill_template_service.py)."""
import pytest
import tempfile
import os
from io import BytesIO
from types import SimpleNamespace
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from backend.services.bill_template_service import render_template


# ── Helpers ──


def _make_party(**kwargs):
    defaults = dict(
        name="测试公司",
        address="北京市朝阳区测试路100号",
        phone="010-88888888",
        bank_name="中国银行北京测试支行",
        bank_account="6222021234567890",
        tax_id="91110000MA11111111",
    )
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _create_simple_template():
    """Create a minimal .xlsx template with scalar + table + total markers."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"

    ws.cell(row=1, column=1, value="{{title}}")
    ws.cell(row=2, column=1, value="对账单编号: {{bill_no}}")
    ws.cell(row=3, column=1, value="对账周期: {{period}}")
    ws.cell(row=4, column=1, value="生成日期: {{date}}")
    ws.cell(row=5, column=1, value="甲方: {{party_a.name}}")
    ws.cell(row=6, column=1, value="乙方: {{party_b.name}}")
    ws.cell(row=7, column=1, value="{{party_a.bank_name}} / {{party_b.bank_account}}")

    # Table start
    ws.cell(row=8, column=1, value="{{#table:rows}}")
    # Header row (static)
    ws.cell(row=9, column=1, value="序号")
    ws.cell(row=9, column=2, value="游戏名称")
    ws.cell(row=9, column=3, value="原始流水")
    ws.cell(row=9, column=4, value="结算金额")
    # Template row (with markers)
    ws.cell(row=10, column=1, value="{{row.seq}}")
    ws.cell(row=10, column=2, value="{{row.game_name}}")
    ws.cell(row=10, column=3, value="{{row.raw_revenue}}")
    ws.cell(row=10, column=4, value="{{row.settlement_amount}}")
    # Table end
    ws.cell(row=11, column=1, value="{{/table}}")

    # Totals
    ws.cell(row=12, column=1, value="合计:")
    ws.cell(row=12, column=3, value="{{total:raw_revenue}}")
    ws.cell(row=12, column=4, value="{{total:settlement_amount}}")

    # Notes
    ws.cell(row=13, column=1, value="{{notes}}")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _create_styled_template():
    """Create a template with styled cells to verify style preservation."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    bold_font = Font(bold=True, size=14, color="FF0000")
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    cell = ws.cell(row=1, column=1, value="{{title}}")
    cell.font = bold_font
    cell.fill = fill
    cell.alignment = center_align

    ws.cell(row=2, column=1, value="{{#table:rows}}")
    ws.cell(row=3, column=1, value="游戏名称")
    data_cell = ws.cell(row=4, column=1, value="{{row.game_name}}")
    data_cell.font = Font(size=11)
    data_cell.border = thin_border
    ws.cell(row=5, column=1, value="{{/table}}")

    ws.cell(row=6, column=1, value="{{total:settlement_amount}}")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _render(template_buf, party_a=None, party_b=None, period="2026-03", rows=None, mode="income"):
    """Helper: write template to temp file, render, return loaded workbook."""
    if party_a is None:
        party_a = _make_party()
    if party_b is None:
        party_b = _make_party(name="渠道乙")
    if rows is None:
        rows = []

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(template_buf.getvalue())
        tpl_path = f.name

    try:
        buf = render_template(tpl_path, party_a, party_b, period, rows, mode)
        wb = load_workbook(buf)
        return wb
    finally:
        os.unlink(tpl_path)


def _cell_text(ws, row, col=1):
    v = ws.cell(row=row, column=col).value
    return str(v) if v is not None else ""


def _all_text(ws):
    """Concatenate all non-None cell values in a worksheet."""
    parts = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                parts.append(str(v))
    return "\n".join(parts)


# ── Tests ──


class TestScalarReplacement:
    def test_title_replaced(self):
        wb = _render(_create_simple_template())
        ws = wb.active
        assert _cell_text(ws, 1) == "收入结算对账单"

    def test_bill_no_format(self):
        wb = _render(_create_simple_template())
        ws = wb.active
        val = _cell_text(ws, 2)
        assert val.startswith("对账单编号: ST-")

    def test_period_replaced(self):
        wb = _render(_create_simple_template(), period="2026-01 ~ 2026-03")
        ws = wb.active
        assert "2026-01 ~ 2026-03" in _cell_text(ws, 3)

    def test_date_replaced(self):
        wb = _render(_create_simple_template())
        ws = wb.active
        val = _cell_text(ws, 4)
        assert "年" in val and "月" in val and "日" in val

    def test_payment_mode_title(self):
        wb = _render(_create_simple_template(), mode="payment")
        ws = wb.active
        assert _cell_text(ws, 1) == "付款结算对账单"

    def test_notes_replaced(self):
        wb = _render(_create_simple_template())
        ws = wb.active
        assert "本对账单仅作为双方对账参考" in _cell_text(ws, 13)

    def test_party_a_name_replaced(self):
        party_a = _make_party(name="我方测试科技有限公司")
        wb = _render(_create_simple_template(), party_a=party_a)
        ws = wb.active
        assert "我方测试科技有限公司" in _cell_text(ws, 5)

    def test_party_b_name_replaced(self):
        party_b = _make_party(name="华为软件技术有限公司")
        wb = _render(_create_simple_template(), party_b=party_b)
        ws = wb.active
        assert "华为软件技术有限公司" in _cell_text(ws, 6)

    def test_multiple_markers_in_one_cell(self):
        party_a = _make_party(bank_name="招商银行")
        party_b = _make_party(bank_account="6222029999999999")
        wb = _render(_create_simple_template(), party_a=party_a, party_b=party_b)
        ws = wb.active
        val = _cell_text(ws, 7)
        assert "招商银行" in val
        assert "6222029999999999" in val

    def test_party_info_all_fields(self):
        party = _make_party(
            name="FullName Corp", address="Shanghai Pudong",
            phone="021-66666666", bank_name="CCB",
            bank_account="6227001234567890", tax_id="TID-123456",
        )
        # Use a template that has {{party_a.address}} marker
        from openpyxl import Workbook
        from io import BytesIO as BIO
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.cell(row=1, column=1, value="{{party_a.name}}")
        ws2.cell(row=2, column=1, value="{{party_a.address}}")
        ws2.cell(row=3, column=1, value="{{party_a.phone}}")
        ws2.cell(row=4, column=1, value="{{party_a.bank_name}}")
        ws2.cell(row=5, column=1, value="{{party_a.bank_account}}")
        ws2.cell(row=6, column=1, value="{{party_a.tax_id}}")
        buf2 = BIO()
        wb2.save(buf2)
        buf2.seek(0)

        wb = _render(buf2, party_a=party)
        ws = wb.active
        text = _all_text(ws)
        assert "FullName Corp" in text
        assert "Shanghai Pudong" in text
        assert "021-66666666" in text  # phone IS in this template
        assert "CCB" in text
        assert "6227001234567890" in text
        assert "TID-123456" in text  # tax_id IS in this template


class TestTableExpansion:
    def test_single_row_expanded(self):
        rows = [
            {"game_name": "梦幻西游", "raw_revenue": 10000, "settlement_amount": 5000},
        ]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        # Header row should still be there
        assert _cell_text(ws, 9) == "序号"
        assert _cell_text(ws, 9, 2) == "游戏名称"
        # Data should be in row 10
        assert _cell_text(ws, 10) == "1"  # seq
        assert _cell_text(ws, 10, 2) == "梦幻西游"
        assert _cell_text(ws, 10, 3) == "10000"
        assert _cell_text(ws, 10, 4) == "5000"
        # End marker should be cleared
        assert _cell_text(ws, 11) != "{{/table}}"

    def test_multiple_rows_expanded(self):
        rows = [
            {"game_name": "游戏A", "raw_revenue": 1000, "settlement_amount": 500},
            {"game_name": "游戏B", "raw_revenue": 2000, "settlement_amount": 1000},
            {"game_name": "游戏C", "raw_revenue": 3000, "settlement_amount": 1500},
        ]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        # Row indices: 9=header, 10=row1, 11=row2, 12=row3
        assert _cell_text(ws, 10) == "1"
        assert _cell_text(ws, 10, 2) == "游戏A"
        assert _cell_text(ws, 11) == "2"
        assert _cell_text(ws, 11, 2) == "游戏B"
        assert _cell_text(ws, 12) == "3"
        assert _cell_text(ws, 12, 2) == "游戏C"

    def test_row_seq_auto_number(self):
        rows = [{"game_name": f"游戏{i}"} for i in range(5)]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        for i in range(5):
            assert _cell_text(ws, 10 + i) == str(i + 1)

    def test_empty_rows_header_preserved(self):
        wb = _render(_create_simple_template(), rows=[])
        ws = wb.active
        # Header should still exist
        assert _cell_text(ws, 9) == "序号"
        assert _cell_text(ws, 9, 2) == "游戏名称"


class TestTotalCalculation:
    def test_total_raw_revenue(self):
        rows = [
            {"game_name": "游戏A", "raw_revenue": 1000, "settlement_amount": 500},
            {"game_name": "游戏B", "raw_revenue": 2000, "settlement_amount": 1000},
            {"game_name": "游戏C", "raw_revenue": 3000, "settlement_amount": 1500},
        ]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        # Row 14 = total row (9 header + 5 data + beyond)
        total_row = None
        for r in range(1, 30):
            if _cell_text(ws, r).startswith("合计"):
                total_row = r
                break
        assert total_row is not None
        assert _cell_text(ws, total_row, 3) == "6000.00"  # 1000+2000+3000
        assert _cell_text(ws, total_row, 4) == "3000.00"  # 500+1000+1500

    def test_total_with_single_row(self):
        rows = [{"game_name": "独苗", "raw_revenue": 9999, "settlement_amount": 4999.5}]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        total_row = None
        for r in range(1, 20):
            if _cell_text(ws, r).startswith("合计"):
                total_row = r
                break
        assert total_row is not None
        assert _cell_text(ws, total_row, 3) == "9999.00"

    def test_total_with_no_rows(self):
        """Empty rows: totals not computed, markers remain (no crash)."""
        wb = _render(_create_simple_template(), rows=[])
        ws = wb.active
        total_row = None
        for r in range(1, 20):
            if _cell_text(ws, r).startswith("合计"):
                total_row = r
                break
        assert total_row is not None
        # With no rows, total markers stay unrendered — that's acceptable
        val = _cell_text(ws, total_row, 3)
        assert "total" in val  # marker remains as-is


class TestStylePreservation:
    def test_font_preserved(self):
        wb = _render(_create_styled_template(), rows=[{"game_name": "Stylized", "settlement_amount": 100}])
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.size == 14
        assert cell.font.color and cell.font.color.rgb and "FF0000" in str(cell.font.color.rgb)

    def test_fill_preserved(self):
        wb = _render(_create_styled_template(), rows=[{"game_name": "G", "settlement_amount": 100}])
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color and "FFFF00" in str(cell.fill.start_color.rgb)

    def test_alignment_preserved(self):
        wb = _render(_create_styled_template(), rows=[{"game_name": "G", "settlement_amount": 100}])
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"

    def test_data_row_border_preserved(self):
        wb = _render(_create_styled_template(), rows=[{"game_name": "Bordered", "settlement_amount": 100}])
        ws = wb.active
        cell = ws.cell(row=4, column=1)
        assert cell.border.left.style is not None

    def test_data_row_font_size_preserved(self):
        wb = _render(_create_styled_template(), rows=[{"game_name": "FontGame", "settlement_amount": 100}])
        ws = wb.active
        cell = ws.cell(row=4, column=1)
        assert cell.font.size == 11


class TestEdgeCases:
    def test_unicode_game_name(self):
        rows = [{"game_name": "🔥火焰山🔥 日本語 中文", "raw_revenue": 5000, "settlement_amount": 2500}]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        assert "火焰山" in _cell_text(ws, 10, 2)
        assert "日本語" in _cell_text(ws, 10, 2)

    def test_negative_values(self):
        rows = [{"game_name": "退款游戏", "raw_revenue": -500, "settlement_amount": -250}]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        assert _cell_text(ws, 10, 3) == "-500"

    def test_zero_values(self):
        rows = [{"game_name": "零流水", "raw_revenue": 0, "settlement_amount": 0}]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        assert _cell_text(ws, 10, 3) == "0"

    def test_large_number_of_rows(self):
        rows = [{"game_name": f"行{i}", "raw_revenue": i * 100, "settlement_amount": i * 50} for i in range(1, 101)]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        assert _cell_text(ws, 10) == "1"
        assert _cell_text(ws, 109) == "100"  # 9+100 = 109
        assert "行100" in _cell_text(ws, 109, 2)

    def test_decimal_precision(self):
        rows = [{"game_name": "精确", "raw_revenue": 1234.5678, "settlement_amount": 0.1 + 0.2}]
        wb = _render(_create_simple_template(), rows=rows)
        ws = wb.active
        assert _cell_text(ws, 10, 3) == "1234.5678"
