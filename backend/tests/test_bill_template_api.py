"""Integration tests for bill template API + export flow."""
import pytest
import os
from io import BytesIO
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from backend.main import app
from backend.database import get_db
from backend import models


# ── Helpers ──


def _make_template_xlsx():
    """Create a simple .xlsx template in memory and return as bytes."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="{{title}}")
    ws.cell(row=2, column=1, value="{{#table:rows}}")
    ws.cell(row=3, column=1, value="game_name")
    ws.cell(row=3, column=2, value="raw_revenue")
    ws.cell(row=3, column=3, value="settlement_amount")
    ws.cell(row=4, column=1, value="{{row.game_name}}")
    ws.cell(row=4, column=2, value="{{row.raw_revenue}}")
    ws.cell(row=4, column=3, value="{{row.settlement_amount}}")
    ws.cell(row=5, column=1, value="{{/table}}")
    ws.cell(row=6, column=1, value="total_revenue: {{total:raw_revenue}}")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.mark.asyncio
async def test_upload_and_export_with_template(db_session):
    """Full flow: upload template -> export with template_id -> verify output."""
    from backend import models

    party_a = models.PartyInfo(
        party_type="our_company", name="CompanyA", address="addr",
        bank_name="CMB", bank_account="acct1", tax_id="tax1",
    )
    party_b = models.PartyInfo(
        party_type="channel", name="ChannelB", address="addr",
        bank_name="CMB_SZ", bank_account="acct2", tax_id="tax2",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    tpl_id = None
    try:
        # Step 1: Upload template
        tpl_bytes = _make_template_xlsx()
        upload_resp = client.post(
            "/api/bill-templates",
            data={
                "name": "ChannelB Template",
                "bill_type": "income",
                "description": "for channel B",
            },
            files={"file": ("template.xlsx", tpl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_resp.status_code == 200, f"Upload failed: {upload_resp.text}"
        tpl_data = upload_resp.json()["data"]
        tpl_id = tpl_data["id"]
        assert tpl_data["name"] == "ChannelB Template"
        assert tpl_data["bill_type"] == "income"

        # Step 2: List templates
        list_resp = client.get("/api/bill-templates?bill_type=income")
        assert list_resp.status_code == 200
        ids = [t["id"] for t in list_resp.json()["data"]]
        assert tpl_id in ids

        # Step 3: Export bill with template
        rows = [
            {"game_name": "GameX", "raw_revenue": 5000, "settlement_amount": 2500},
            {"game_name": "GameY", "raw_revenue": 8000, "settlement_amount": 4000},
        ]
        export_resp = client.post("/api/settlement/bill", json={
            "mode": "income",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-03",
            "end_month": "2026-03",
            "rows": rows,
            "template_id": tpl_id,
        })
        assert export_resp.status_code == 200, f"Export failed: {export_resp.text}"
        assert export_resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Step 4: Verify Excel content
        wb = load_workbook(BytesIO(export_resp.content))
        ws = wb.active

        all_vals = []
        for r in range(1, 20):
            for c in range(1, 10):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_vals.append(str(v))
        text = "\n".join(all_vals)

        # Template markers should be replaced
        assert "收入结算对账单" in text, f"Title not replaced. Content:\n{text}"
        assert "GameX" in text, f"Row data missing. Content:\n{text}"
        assert "GameY" in text
        assert "5000" in text
        assert "8000" in text
        # Total should be calculated
        assert "13000.00" in text, f"Total not calculated. Content:\n{text}"
        # No raw markers should remain
        assert "{{" not in text, f"Unrendered markers remain. Content:\n{text}"

    finally:
        app.dependency_overrides.clear()
        if tpl_id:
            try:
                client.delete(f"/api/bill-templates/{tpl_id}")
            except Exception:
                pass


@pytest.mark.asyncio
async def test_export_without_template_backward_compat(db_session):
    """Without template_id -> falls back to hardcoded generation (no regression)."""
    from backend import models

    party_a = models.PartyInfo(
        party_type="our_company", name="CompanyA", address="addr",
        bank_name="BankA", bank_account="acctA", tax_id="taxA",
    )
    party_b = models.PartyInfo(
        party_type="channel", name="ChannelB", address="addr",
        bank_name="BankB", bank_account="acctB", tax_id="taxB",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    try:
        rows = [
            {"game_id": "G001", "game_name": "Game1", "raw_revenue": 1000, "real_revenue": 900,
             "vouchers": 50, "test": 20, "welfare": 10, "bad_debt": 5, "total_deductions": 85,
             "split_rate": 0.5, "channel_fee_rate": 0.1, "tax_rate": 0.05,
             "settlement_amount": 348.84},
        ]
        resp = client.post("/api/settlement/bill", json={
            "mode": "income",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-03",
            "end_month": "2026-03",
            "rows": rows,
        })
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        all_vals = []
        for r in range(1, 50):
            for c in range(1, 18):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_vals.append(str(v))
        text = "|".join(all_vals)
        assert "CompanyA" in text
        assert "G001" in text
        assert "合  计" in text
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_export_with_invalid_template_id(db_session):
    """Non-existent template_id -> 404."""
    from backend import models

    party_a = models.PartyInfo(
        party_type="our_company", name="A", address="addr",
        bank_name="B", bank_account="acct", tax_id="t",
    )
    party_b = models.PartyInfo(
        party_type="channel", name="B", address="addr",
        bank_name="C", bank_account="acct", tax_id="t",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    try:
        resp = client.post("/api/settlement/bill", json={
            "mode": "income",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-03",
            "end_month": "2026-03",
            "rows": [],
            "template_id": 99999,
        })
        assert resp.status_code == 404
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_template_download(db_session):
    """Upload template -> download it back -> content matches."""
    from backend import models

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    tpl_id = None
    try:
        tpl_bytes = _make_template_xlsx()
        upload_resp = client.post(
            "/api/bill-templates",
            data={"name": "DownloadTest", "bill_type": "all"},
            files={"file": ("test.xlsx", tpl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_resp.status_code == 200
        tpl_id = upload_resp.json()["data"]["id"]

        download_resp = client.get(f"/api/bill-templates/{tpl_id}/download")
        assert download_resp.status_code == 200
        assert download_resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert len(download_resp.content) > 0

        wb = load_workbook(BytesIO(download_resp.content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "{{title}}"
        assert ws.cell(row=2, column=1).value == "{{#table:rows}}"
    finally:
        app.dependency_overrides.clear()
        if tpl_id:
            try:
                client.delete(f"/api/bill-templates/{tpl_id}")
            except Exception:
                pass


@pytest.mark.asyncio
async def test_template_delete(db_session):
    """Upload -> delete -> template no longer listed."""
    from backend import models

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    tpl_id = None
    try:
        tpl_bytes = _make_template_xlsx()
        upload_resp = client.post(
            "/api/bill-templates",
            data={"name": "DelTest", "bill_type": "payment"},
            files={"file": ("del.xlsx", tpl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_resp.status_code == 200
        tpl_id = upload_resp.json()["data"]["id"]

        del_resp = client.delete(f"/api/bill-templates/{tpl_id}")
        assert del_resp.status_code == 200

        list_resp = client.get("/api/bill-templates")
        ids = [t["id"] for t in list_resp.json()["data"]]
        assert tpl_id not in ids
        tpl_id = None
    finally:
        app.dependency_overrides.clear()
        if tpl_id:
            try:
                client.delete(f"/api/bill-templates/{tpl_id}")
            except Exception:
                pass


@pytest.mark.asyncio
async def test_export_template_with_payment_mode(db_session):
    """Payment mode + template -> correct title."""
    from backend import models

    party_a = models.PartyInfo(
        party_type="our_company", name="A", address="addr",
        bank_name="B", bank_account="acct", tax_id="t",
    )
    party_b = models.PartyInfo(
        party_type="publisher", name="PubB", address="addr",
        bank_name="C", bank_account="acct", tax_id="t",
    )
    db_session.add_all([party_a, party_b])
    await db_session.commit()
    await db_session.refresh(party_a)
    await db_session.refresh(party_b)

    async def _override():
        yield db_session
    app.dependency_overrides[get_db] = _override

    client = TestClient(app, raise_server_exceptions=False, client=("127.0.0.1", 50000))
    tpl_id = None
    try:
        tpl_bytes = _make_template_xlsx()
        upload_resp = client.post(
            "/api/bill-templates",
            data={"name": "Universal", "bill_type": "all"},
            files={"file": ("t.xlsx", tpl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_resp.status_code == 200
        tpl_id = upload_resp.json()["data"]["id"]

        rows = [{"game_name": "Game", "raw_revenue": 10000, "settlement_amount": 5000}]
        resp = client.post("/api/settlement/bill", json={
            "mode": "payment",
            "party_id_a": party_a.id,
            "party_id_b": party_b.id,
            "start_month": "2026-04",
            "end_month": "2026-04",
            "rows": rows,
            "template_id": tpl_id,
        })
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        text = str(ws.cell(row=1, column=1).value or "")
        assert "付款结算对账单" in text
    finally:
        app.dependency_overrides.clear()
        if tpl_id:
            try:
                client.delete(f"/api/bill-templates/{tpl_id}")
            except Exception:
                pass
