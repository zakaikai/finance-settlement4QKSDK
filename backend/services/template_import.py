# -*- coding: utf-8 -*-
"""Template-based Excel import pipeline: parse → FK resolve → validate → conflict check → import."""
from decimal import Decimal
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession

from .. import models
from . import fk_resolver
from .template_defs import TEMPLATE_DEFS


async def parse_excel(file_path, template_type):
    td = TEMPLATE_DEFS.get(template_type)
    if td is None:
        return {"error": f"Unknown template type: {template_type}"}

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    col_map = {}
    rows = []
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i == 0:
            for idx, val in enumerate(row):
                if val is not None:
                    col_map[str(val).strip()] = idx
            missing = [c for c in td["columns"] if c not in col_map]
            if missing:
                wb.close()
                return {"error": f"Missing columns: {', '.join(missing)}"}
            continue

        if all(v is None for v in row):
            continue

        record = {}
        row_errors = []
        for ci, col_name in enumerate(td["columns"]):
            raw = row[col_map[col_name]] if col_map[col_name] < len(row) else None
            cell_type = td["types"][ci]
            required = td["required"][ci]

            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                if required:
                    row_errors.append(f"{col_name} is required")
                record[col_name] = None
                continue

            if cell_type == Decimal:
                try:
                    record[col_name] = Decimal(str(raw))
                except Exception:
                    row_errors.append(f"{col_name} must be a number")
            elif cell_type == str:
                record[col_name] = str(raw).strip()
            else:
                record[col_name] = raw

        if row_errors:
            errors.append({"row": i + 1, "errors": row_errors})
        else:
            rows.append(record)

    wb.close()
    return {
        "total_rows": len(rows),
        "preview_rows": rows[:5],
        "rows": rows,
        "errors": errors,
        "has_conflict": False,
        "conflict_count": 0,
    }


async def resolve_foreign_keys(session, template_type, rows):
    td = TEMPLATE_DEFS.get(template_type)
    if td is None:
        return {"error": f"Unknown template type: {template_type}"}

    errors = []
    await fk_resolver.reset()

    for ri, record in enumerate(rows):
        row_num = ri + 2

        for fr in td.get("fk_resolves", []):
            fk_val = await fk_resolver.resolve(
                session, fr["fk_model"], fr["name_col"],
                record.get(fr["name_field"]), fr["cache_key"],
            )
            if fk_val is None:
                errors.append({"row": row_num, "error": fr["error_msg"]})
            record[fr["target_field"]] = fk_val

        for fld in td.get("date_fields", []):
            v = record.get(fld)
            if v and isinstance(v, str):
                try:
                    record[fld] = datetime.strptime(v, "%Y-%m-%d").date()
                except ValueError:
                    errors.append({"row": row_num, "error": f"Invalid date format for {fld}"})

        if td.get("raw_txn"):
            bk_id = record.get("backend_channel_id")
            sub_name = record.get("sub_channel_name")
            if sub_name and bk_id is not None:
                ck = ("sub_channels_v2", sub_name, bk_id)
                async def _sub_query():
                    stmt = select(models.SubChannel.sub_channel_id).where(
                        models.SubChannel.sub_channel_name == sub_name,
                        models.SubChannel.backend_channel_id == bk_id,
                    )
                    row = (await session.execute(stmt)).one_or_none()
                    return row[0] if row else None
                sub_id = await fk_resolver.resolve_raw(ck, _sub_query)
                if sub_id is None:
                    errors.append({"row": row_num, "error": "Sub channel not found"})
                record["sub_channel_id"] = sub_id
            elif sub_name is None:
                errors.append({"row": row_num, "error": "sub_channel_name is required"})

            date_val = record.get("record_date")
            if date_val:
                try:
                    if isinstance(date_val, str):
                        dt = datetime.strptime(date_val, "%Y-%m-%d")
                        record["record_date"] = dt.date()
                        record["month"] = dt.strftime("%Y-%m")
                except ValueError:
                    errors.append({"row": row_num, "error": "Invalid date format"})

        game_id = record.get("game_id")
        if game_id is not None and template_type != "games":
            exists = await session.execute(
                select(models.Game.game_id).where(models.Game.game_id == game_id)
            )
            if exists.scalar_one_or_none() is None:
                errors.append({"row": row_num, "error": f"Game '{game_id}' not found in games table"})

    return {"errors": errors}


async def validate_values(template_type, rows):
    errors = []
    td = TEMPLATE_DEFS.get(template_type)
    for ri, record in enumerate(rows):
        row_num = ri + 2
        for ci, col_name in enumerate(td["columns"]):
            val = record.get(col_name)
            if val is None:
                continue
            if td["types"][ci] == Decimal and val < 0:
                errors.append({"row": row_num, "error": f"{col_name} cannot be negative"})
            if td["types"][ci] == Decimal:
                exp = val.as_tuple().exponent
                if exp < -4:
                    errors.append({"row": row_num, "error": f"{col_name} has too many decimal places (max 4)"})
            if td["types"][ci] == str and isinstance(val, str) and len(val) > 200:
                errors.append({"row": row_num, "error": f"{col_name} exceeds 200 chars"})

        if template_type == "games":
            dr = record.get("discount_rate")
            if dr is not None and (dr < 0 or dr > 1):
                errors.append({"row": row_num, "error": "discount_rate must be between 0 and 1"})
        elif template_type in ("income_split", "payment_split"):
            for field in ("split_rate", "channel_fee_rate", "tax_rate"):
                v = record.get(field)
                if v is not None and (v < 0 or v > 1):
                    errors.append({"row": row_num, "error": f"{field} must be between 0 and 1"})
        if template_type == "raw_transactions":
            amt = record.get("amount")
            if amt is not None and amt <= 0:
                errors.append({"row": row_num, "error": "amount must be positive"})
            rd = record.get("record_date")
            if rd is not None and isinstance(rd, str):
                parts = rd.split("-")
                if not (len(parts) == 3 and len(parts[0]) == 4 and len(parts[1]) == 2 and len(parts[2]) == 2
                        and parts[0].isdigit() and parts[1].isdigit() and parts[2].isdigit()):
                    errors.append({"row": row_num, "error": f"Invalid record_date format: {rd} (expected YYYY-MM-DD)"})

        month_val = record.get("month")
        if month_val is not None and isinstance(month_val, str):
            parts = month_val.split("-")
            if not (len(parts) == 2 and len(parts[0]) == 4 and len(parts[1]) == 2
                    and parts[0].isdigit() and parts[1].isdigit()
                    and 1 <= int(parts[1]) <= 12):
                errors.append({"row": row_num, "error": f"Invalid month format: {month_val} (expected YYYY-MM)"})

    batch_key = td.get("batch_key_fields") or td["unique_fields"]
    if batch_key:
        seen = {}
        for ri, record in enumerate(rows):
            row_num = ri + 2
            key = tuple(record.get(f) for f in batch_key)
            if key in seen:
                dup_fields = ", ".join(f"{f}={key[i]}" for i, f in enumerate(batch_key))
                errors.append({"row": row_num, "error": f"Duplicate {dup_fields} (also in row {seen[key]})"})
            else:
                seen[key] = row_num

    return {"errors": errors}


async def check_conflicts(session, template_type, rows):
    td = TEMPLATE_DEFS.get(template_type)
    if td is None or td["unique_fields"] is None:
        return {"has_conflict": False, "conflict_count": 0}
    conflict_count = 0
    for record in rows:
        filters = {}
        for uf in td["unique_fields"]:
            val = record.get(uf)
            if val is not None:
                filters[uf] = val
        if not filters:
            continue
        existing = (await session.execute(select(td["model"]).filter_by(**filters))).scalar_one_or_none()
        if existing is not None:
            conflict_count += 1
    return {"has_conflict": conflict_count > 0, "conflict_count": conflict_count}


async def import_channels(session, rows) -> int:
    """Upsert 3-level channel hierarchy: ChannelCategory → BackendChannel → SubChannel.

    Delegates to basic_data_service.batch_save_channels (shared 3-level logic).
    """
    from types import SimpleNamespace
    from .basic_data_service import batch_save_channels

    items = [
        SimpleNamespace(
            action="create",
            channel_name=r["channel_name"],
            backend_channel_name=r["backend_channel_name"],
            sub_channel_name=r["sub_channel_name"],
        )
        for r in rows
    ]
    await batch_save_channels(session, items, skip_duplicates=True)
    return len(rows)


async def import_raw_transactions(session, rows, overwrite=False) -> int:
    """导入原始流水 → 直接聚合写入 raw_settlements。

    流程：
    1. 解析 sub_channel_id → 一级 channel_id（三级渠道 at import time only）
    2. 按 (channel_id, game_id, month) 聚合 raw_revenue
    3. UPSERT 写入 raw_settlements（仅存聚合行，不存行级数据）
    """
    from collections import defaultdict
    from datetime import datetime

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    imported = 0

    # 1. Resolve sub_channel_id if present
    for record in rows:
        if record.get("sub_channel_id") and not record.get("channel_id"):
            ch_rows = (await session.execute(
                select(models.SubChannel.sub_channel_id, models.BackendChannel.channel_id)
                .join(models.BackendChannel, models.SubChannel.backend_channel_id == models.BackendChannel.backend_channel_id)
                .where(models.SubChannel.sub_channel_id == record["sub_channel_id"])
            )).all()
            if ch_rows:
                record["channel_id"] = ch_rows[0][1]

    # 2. Aggregate by (channel_id, game_id, month)
    agg: dict[tuple, Decimal] = {}
    for record in rows:
        ch_id = record.get("channel_id")
        g_id = record.get("game_id")
        m = record.get("month")
        amount = record.get("amount", 0)
        if ch_id and g_id and m:
            key = (ch_id, g_id, m)
            agg[key] = agg.get(key, Decimal("0")) + Decimal(str(amount))

    if not agg:
        return 0

    # 3. Get game names + channel names (batch prefetch)
    gids = list({k[1] for k in agg})
    game_rows = (await session.execute(
        select(models.Game.game_id, models.Game.game_name).where(models.Game.game_id.in_(gids))
    )).all()
    game_names = {r.game_id: r.game_name for r in game_rows}

    ch_ids = list({k[0] for k in agg})
    ch_rows = (await session.execute(
        select(models.ChannelCategory.channel_id, models.ChannelCategory.channel_name)
        .where(models.ChannelCategory.channel_id.in_(ch_ids))
    )).all()
    channel_names = {r.channel_id: r.channel_name for r in ch_rows}

    # 4. Overwrite mode: clear all rows for affected months (all channels)
    if overwrite:
        affected_months = list({k[2] for k in agg})
        logger = __import__('logging').getLogger(__name__)
        logger.warning(f"[raw-import] overwrite=True, clearing {len(affected_months)} months: {affected_months}")
        for m in affected_months:
            result = await session.execute(
                delete(models.RawSettlement).where(
                    models.RawSettlement.month == m,
                )
            )
            logger.warning(f"[raw-import] deleted month={m}: {result.rowcount} rows")

    # 5. Write aggregated rows
    for (ch_id, g_id, m), raw_rev in agg.items():
        ch_name = channel_names.get(ch_id, "")
        if overwrite:
            # Fresh insert — old rows already cleared
            session.add(models.RawSettlement(
                channel_id=ch_id, channel_name=ch_name,
                game_id=g_id, game_name=game_names.get(g_id, ""),
                month=m, raw_revenue=raw_rev,
                created_at=now, updated_at=now,
            ))
        else:
            # Accumulate mode: UPSERT with +=
            row = (await session.execute(
                select(models.RawSettlement).where(
                    models.RawSettlement.channel_id == ch_id,
                    models.RawSettlement.game_id == g_id,
                    models.RawSettlement.month == m,
                )
            )).scalar_one_or_none()
            if row:
                row.raw_revenue += raw_rev
                row.channel_name = ch_name
                row.game_name = game_names.get(g_id, "")
                row.updated_at = now
            else:
                session.add(models.RawSettlement(
                    channel_id=ch_id, channel_name=ch_name,
                    game_id=g_id, game_name=game_names.get(g_id, ""),
                    month=m, raw_revenue=raw_rev,
                    created_at=now, updated_at=now,
                ))
        imported += 1

    await session.commit()
    return imported


async def import_generic(session, template_type, rows, overwrite=False) -> int:
    """Generic import for simple entities: unique-key upsert or overwrite."""
    td = TEMPLATE_DEFS.get(template_type)
    model_class = td["model"]
    unique_fields = td["unique_fields"]

    imported = 0
    for record in rows:
        model_columns = {k: v for k, v in record.items() if hasattr(model_class, k) and v is not None}
        if overwrite and unique_fields:
            filters = {uf: record[uf] for uf in unique_fields if uf in record}
            if filters:
                await session.execute(delete(model_class).filter_by(**filters))
        if not overwrite and unique_fields:
            filters = {uf: record[uf] for uf in unique_fields if uf in record}
            if filters:
                existing = (await session.execute(select(model_class).filter_by(**filters))).scalar_one_or_none()
                if existing is not None:
                    continue
        instance = model_class(**model_columns)
        session.add(instance)
        imported += 1

    await session.commit()
    return imported


async def import_data(session, template_type, rows, overwrite=False):
    """Dispatch to the correct import strategy based on template type."""
    if template_type == "channels":
        return {"imported": await import_channels(session, rows)}
    if template_type == "raw_transactions":
        return {"imported": await import_raw_transactions(session, rows, overwrite)}
    td = TEMPLATE_DEFS.get(template_type)
    if td is None:
        return {"error": f"Unknown template type: {template_type}"}
    return {"imported": await import_generic(session, template_type, rows, overwrite)}
