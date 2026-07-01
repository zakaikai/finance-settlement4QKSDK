"""Excel template rendering engine for bill export.

Supports markers:
  {{field}}              — single value replacement
  {{party_a.field}}      — party info (name, address, phone, bank_name, etc.)
  {{party_b.field}}      — same for party b
  {{#table:name}}        — table region start
  {{/table}}             — table region end
  {{row.field}}          — per-row field inside table region
  {{total:field}}        — column sum below table region
"""
import re
import os
from copy import copy
from datetime import datetime, date
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ── Marker patterns ──

_TABLE_START_RE = re.compile(r'\{\{#table:(\w+)\}\}')
_TABLE_END_RE = re.compile(r'\{\{/table\}\}')
_ROW_MARKER_RE = re.compile(r'\{\{row\.([^}]+)\}\}')
_TOTAL_MARKER_RE = re.compile(r'\{\{total:(\w+)\}\}')
_SCALAR_MARKER_RE = re.compile(r'\{\{([a-zA-Z_][\w.]*)\}\}')

# Matches cell references in formulas:  A1, $A1, A$1, $A$1, Sheet2!A1, etc.
# Groups: 1=$-prefix-col, 2=column, 3=$-prefix-row, 4=row-number
_CELL_REF_RE = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

# ── Rendering ──


def render_template(template_path: str, party_a, party_b, period: str,
                    rows: list[dict], mode: str) -> BytesIO:
    """Render an Excel template with settlement data.

    Parameters
    ----------
    template_path : str — path to the .xlsx template file
    party_a, party_b — PartyInfo ORM objects (or None)
    period : str — billing period display string
    rows : list[dict] — settlement data rows from the frontend
    mode : str — "income" or "payment"

    Returns BytesIO of the rendered xlsx.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    ctx = _build_context(party_a, party_b, period, mode)

    # Phase 1: expand table regions (bottom-up to preserve row numbers)
    regions = _find_table_regions(ws)
    for region in reversed(regions):
        _expand_table(ws, region, rows)

    # Phase 2: replace scalar markers
    _replace_scalars(ws, ctx)

    # Phase 3: replace {{total:field}} markers
    _replace_totals(ws, rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Context builder ──


def _build_context(party_a, party_b, period: str, mode: str) -> dict:
    today = date.today()
    bill_no = f"ST-{today.strftime('%Y%m%d')}-{datetime.now().strftime('%H%M')}"
    title = "收入结算对账单" if mode == "income" else "付款结算对账单"

    ctx = {
        "title": title,
        "bill_no": bill_no,
        "period": period,
        "date": today.strftime("%Y年%m月%d日"),
        "notes": "1. 本对账单仅作为双方对账参考，不作为最终结算凭证。\n"
                 "2. 如有异议，请于收到本对账单后 5 个工作日内联系我方。",
    }

    for prefix, party in [("party_a", party_a), ("party_b", party_b)]:
        if party:
            ctx[f"{prefix}.name"] = party.name or ""
            ctx[f"{prefix}.address"] = party.address or ""
            ctx[f"{prefix}.phone"] = party.phone or ""
            ctx[f"{prefix}.bank_name"] = party.bank_name or ""
            ctx[f"{prefix}.bank_account"] = party.bank_account or ""
            ctx[f"{prefix}.tax_id"] = party.tax_id or ""
        else:
            for attr in ("name", "address", "phone", "bank_name", "bank_account", "tax_id"):
                ctx[f"{prefix}.{attr}"] = ""

    return ctx


# ── Table region detection ──


def _cell_str(cell) -> str:
    return str(cell.value) if cell.value is not None else ""


def _find_table_regions(ws):
    """Scan active sheet and locate all {{#table:}}…{{/table}} regions."""
    regions = []
    start_row = None
    start_name = None

    for row_idx in range(1, ws.max_row + 1):
        for cell in ws[row_idx]:
            val = _cell_str(cell)
            m = _TABLE_START_RE.search(val)
            if m:
                start_row = row_idx
                start_name = m.group(1)
                break
            if _TABLE_END_RE.search(val) and start_row is not None:
                region = _build_region(ws, start_row, row_idx, start_name)
                regions.append(region)
                start_row = None
                start_name = None
                break

    return regions


def _build_region(ws, start_row, end_row, name):
    """Collect static and template rows within a region."""
    static_rows = []
    template_rows = []

    for r in range(start_row + 1, end_row):
        has_row_marker = any(_ROW_MARKER_RE.search(_cell_str(c)) for c in ws[r])
        if has_row_marker:
            template_rows.append(r)
        else:
            static_rows.append(r)

    return {
        "start_row": start_row,
        "end_row": end_row,
        "name": name,
        "static_rows": static_rows,
        "template_rows": template_rows,
    }


# ── Table expansion ──


def _expand_table(ws, region, rows):
    """Expand a table region: duplicate template rows, fill data."""
    if not rows or not region["template_rows"]:
        return

    n_data = len(rows)
    n_tpl = len(region["template_rows"])

    # ── Snapshot original row data from every row in the region ──
    def _snap_row(r):
        snap = []
        for cell in ws[r]:
            snap.append({
                "col": cell.column,
                "value": cell.value,
                "font": copy(cell.font) if cell.has_style else None,
                "fill": copy(cell.fill) if cell.has_style else None,
                "border": copy(cell.border) if cell.has_style else None,
                "alignment": copy(cell.alignment) if cell.has_style else None,
                "number_format": copy(cell.number_format) if cell.has_style else None,
            })
        return snap

    row_snapshots = {r: _snap_row(r) for r in
                     range(region["start_row"], region["end_row"] + 1)}

    # ── Insert rows at the end marker, pushing it down ──
    extra = n_data * n_tpl - n_tpl  # total template slots we need
    if extra > 0:
        insert_at = region["end_row"]

        # Preserve merged cell ranges below insertion point
        # (openpyxl insert_rows shifts cell values but NOT merged ranges)
        shifted_merges = []
        for mcr in list(ws.merged_cells.ranges):
            if mcr.min_row >= insert_at:
                shifted_merges.append(mcr)
                ws.unmerge_cells(str(mcr))

        ws.insert_rows(insert_at, extra)

        # Re-apply merged ranges at shifted positions
        for mcr in shifted_merges:
            ws.merge_cells(
                start_row=mcr.min_row + extra,
                start_column=mcr.min_col,
                end_row=mcr.max_row + extra,
                end_column=mcr.max_col,
            )

    # ── Write output ──
    # Static rows before the first template row are written first,
    # then data rows (template repeated), then static rows after the last template row.
    first_tpl = min(region["template_rows"])
    last_tpl = max(region["template_rows"])
    static_before = sorted(r for r in region["static_rows"] if r < first_tpl)
    static_after = sorted(r for r in region["static_rows"] if r > last_tpl)

    out_row = region["start_row"]

    # Clear start marker
    _clear_row_cells(ws, out_row)
    out_row += 1

    # Write static rows before template (once)
    for sr in static_before:
        _write_snapshot(ws, out_row, row_snapshots[sr])
        _copy_row_height(ws, sr, out_row)
        out_row += 1

    # Write data rows (template rows repeated for each data row)
    for data_idx, data_row in enumerate(rows):
        for tr in region["template_rows"]:
            row_offset = out_row - tr
            snap = row_snapshots[tr]
            for cell_data in snap:
                dst = ws.cell(row=out_row, column=cell_data["col"])
                dst.value = _replace_row_markers(cell_data["value"], data_row, data_idx)
                if isinstance(dst.value, str) and dst.value.startswith("="):
                    dst.value = _adjust_formula_refs(dst.value, row_offset)
                if cell_data["font"] is not None:
                    dst.font = cell_data["font"]
                if cell_data["fill"] is not None:
                    dst.fill = cell_data["fill"]
                if cell_data["border"] is not None:
                    dst.border = cell_data["border"]
                if cell_data["alignment"] is not None:
                    dst.alignment = cell_data["alignment"]
                if cell_data["number_format"] is not None:
                    dst.number_format = cell_data["number_format"]
            _copy_row_height(ws, tr, out_row)
            out_row += 1

    # Write static rows after template (once, after all data rows)
    for sr in static_after:
        _write_snapshot(ws, out_row, row_snapshots[sr])
        _copy_row_height(ws, sr, out_row)
        out_row += 1

    # Clear the end marker (its position = start + 1 + static_before_count + n_data * n_tpl + static_after_count)
    end_pos = region["start_row"] + 1 + len(static_before) + n_data * n_tpl + len(static_after)
    _clear_row_cells(ws, end_pos)

    # Adjust formulas below the table: expand range ends that referenced the
    # old last template row to now cover all data rows.
    if extra > 0:
        old_last_tpl = max(region["template_rows"])
        new_last_data_row = region["start_row"] + 1 + len(static_before) + n_data * n_tpl - 1
        if old_last_tpl != new_last_data_row:  # range actually expanded
            _expand_formulas_below(ws, old_last_tpl, new_last_data_row, end_pos + 1)


def _write_snapshot(ws, row_num, snap):
    """Write a snapshot (list of cell_data dicts) to a worksheet row."""
    for cell_data in snap:
        dst = ws.cell(row=row_num, column=cell_data["col"])
        dst.value = cell_data["value"]
        if cell_data["font"] is not None:
            dst.font = cell_data["font"]
        if cell_data["fill"] is not None:
            dst.fill = cell_data["fill"]
        if cell_data["border"] is not None:
            dst.border = cell_data["border"]
        if cell_data["alignment"] is not None:
            dst.alignment = cell_data["alignment"]
        if cell_data["number_format"] is not None:
            dst.number_format = cell_data["number_format"]


def _adjust_formula_refs(formula: str, row_offset: int) -> str:
    """Adjust relative row references in an Excel formula by row_offset.

    Only relative rows (no $ prefix) are adjusted; absolute rows ($1) are kept.
    """
    if not formula.startswith("="):
        return formula

    def _adjust(m):
        dollar_col, col, dollar_row, row = m.groups()
        if not dollar_row:
            return f"{dollar_col}{col}{dollar_row}{int(row) + row_offset}"
        return m.group(0)

    return _CELL_REF_RE.sub(_adjust, formula)


def _replace_row_markers(value, data_row, idx):
    """Replace {{row.field}} markers in a cell value with actual data.

    Single-marker cells (e.g. {{row.raw_revenue}}) return raw value so
    Excel number_format (#,##0.00, 0.00%) applies correctly.
    Embedded markers (e.g. "合计：{{row.total}}") return string.
    """
    if value is None:
        return None
    if "{{row." not in str(value):
        return value

    val = str(value).strip()

    # Single marker → raw value (number types preserved for formatting)
    if val.startswith("{{row.") and val.endswith("}}"):
        m = _ROW_MARKER_RE.search(val)
        if m and m.group(0) == val:
            field = m.group(1)
            if field == "seq":
                return idx + 1
            return data_row.get(field)

    # Embedded markers → string replacement
    result = str(value)
    for m in _ROW_MARKER_RE.finditer(result):
        field = m.group(1)
        if field == "seq":
            repl = str(idx + 1)
        else:
            raw = data_row.get(field)
            repl = str(raw) if raw is not None else ""
        result = result.replace(m.group(0), repl)
    return result


# ── Cell / row helpers ──


def _clear_row_cells(ws, row):
    for cell in ws[row]:
        cell.value = None


def _copy_cell_value_and_style(src, dst):
    dst.value = src.value
    _copy_cell_style(src, dst)


def _copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = copy(src.number_format)


def _copy_row_height(ws, src_row, dst_row):
    h = ws.row_dimensions[src_row].height
    if h:
        ws.row_dimensions[dst_row].height = h


# ── Scalar replacement ──


def _replace_scalars(ws, ctx):
    """Replace {{field}} / {{party_a.field}} / {{party_b.field}} markers."""
    for row in ws.iter_rows():
        for cell in row:
            val = _cell_str(cell)
            if "{{" not in val:
                continue
            new_val = val
            for key, value in ctx.items():
                marker = "{{" + key + "}}"
                if marker in new_val:
                    new_val = new_val.replace(marker, str(value))
            if new_val != val:
                cell.value = new_val


def _replace_totals(ws, rows):
    """Replace {{total:field}} markers with summed values."""
    if not rows:
        return
    for row in ws.iter_rows():
        for cell in row:
            val = _cell_str(cell)
            if "{{total:" not in val:
                continue
            new_val = val
            for m in _TOTAL_MARKER_RE.finditer(val):
                field = m.group(1)
                total = sum(
                    float(r.get(field, 0) or 0) for r in rows
                )
                new_val = new_val.replace(m.group(0), f"{total:.2f}")
            if new_val != val:
                cell.value = new_val


# ── Formula adjustment after table expansion ──

_CELL_RANGE_RE = re.compile(
    r'(\$?[A-Z]{1,3})(\$?)(\d+):(\$?[A-Z]{1,3})(\$?)(\d+)'
)


def _expand_formulas_below(ws, old_last_row, new_last_row, start_scan):
    """Expand cell range ends in formulas below an expanded table.

    After inserting rows into a table, SUM/SUBTOTAL formulas below the table
    that reference the original template row range need their range end
    extended to cover the new data rows.

    E.g.  =SUM(J7:J7)  →  =SUM(J7:J11)   (when new_last_row=11)
    """
    for r in range(start_scan, ws.max_row + 1):
        for cell in ws[r]:
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue
            new_val = val
            for m in _CELL_RANGE_RE.finditer(val):
                col1, d1, r1, col2, d2, r2 = m.groups()
                if int(r2) == old_last_row:
                    new_ref = f"{col1}{d1}{r1}:{col2}{d2}{new_last_row}"
                    new_val = new_val.replace(m.group(0), new_ref)
            if new_val != val:
                cell.value = new_val
