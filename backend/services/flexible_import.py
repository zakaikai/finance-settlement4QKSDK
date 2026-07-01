# -*- coding: utf-8 -*-
"""Flexible import: column inference, Excel parse, game name fuzzy match, DB write."""
import logging
from decimal import Decimal
from openpyxl import load_workbook
from sqlalchemy import select, tuple_, or_
from sqlalchemy.ext.asyncio import AsyncSession

from .. import models
from .ocr_service import match_game_names as _fuzzy_match_games
from .field_definitions import FIELD_DEFS

logger = logging.getLogger("finance-settlement")
from .template_defs import FLEXIBLE_FIELD_DEFS
from .split_config_service import save_income_split_config
from .settlement_formula import hydrate_formula_input, compute as compute_settlement
from ..utils.dates import month_bounds


async def parse_flexible_excel(
    file_path: str,
    column_mapping: dict[str, str],
    header_row: int = 1,
) -> dict:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    errors = []

    col_map = {}
    field_defs = {f[0]: f for f in FLEXIBLE_FIELD_DEFS}
    for ci_str, field_key in column_mapping.items():
        if field_key and field_key != "ignore":
            col_map[int(ci_str)] = field_key

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i + 1 < header_row:
            continue
        if i + 1 == header_row:
            continue

        if all(v is None for v in row):
            continue

        record = {}
        row_errors = []
        for ci, field_key in col_map.items():
            raw = row[ci] if ci < len(row) else None
            fdef = field_defs.get(field_key)
            is_money = fdef[2] if fdef else False

            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                record[field_key] = None
                continue

            if is_money:
                try:
                    s = str(raw).replace(",", "").replace("¥", "").replace("￥", "").replace("元", "").strip()
                    record[field_key] = Decimal(s)
                except Exception:
                    row_errors.append(f"列{ci+1} '{raw}' 无法转为数字")
                    record[field_key] = None
            else:
                record[field_key] = str(raw).strip()

        if row_errors:
            errors.append({"row": i + 1, "errors": row_errors})
        elif record:
            rows.append(record)

    wb.close()
    return {"rows": rows, "errors": errors, "total_rows": len(rows)}


async def resolve_flexible_game_names(
    db: AsyncSession,
    rows: list[dict],
    game_name_col: str,
    channel_id: int | None = None,
    month: str | None = None,
) -> tuple[list[dict], list[dict]]:
    names = []
    for r in rows:
        v = r.get(game_name_col, "")
        names.append(str(v) if v else "")

    match_results = await _fuzzy_match_games(db, names)

    # ── Filter: only allow game_ids that exist in raw_settlements for this channel ──
    valid_gids: set[str] = set()
    if channel_id:
        rs_filters = [models.RawSettlement.channel_id == channel_id]
        if month:
            rs_filters.append(models.RawSettlement.month == month)
        rs_rows = (await db.execute(
            select(models.RawSettlement.game_id).where(*rs_filters).distinct()
        )).all()
        valid_gids = {r.game_id for r in rs_rows}

    errors = []

    for i, m in enumerate(match_results):
        row_num = i + 2
        if m["status"] == "none":
            errors.append({"row": row_num, "error": f"游戏名 '{m['candidate']}' 为空"})
            rows[i]["game_id"] = None
            continue

        matched_gid = m["matched_game_id"]

        # If channel filter active and matched game not in raw_settlements, try alternatives
        if valid_gids and matched_gid not in valid_gids:
            # Search candidates for a game that IS in raw_settlements
            candidates = m.get("candidates") or []
            found = False
            for alt in candidates:
                alt_gid = alt.get("game_id", "")
                if alt_gid in valid_gids:
                    matched_gid = alt_gid
                    m["matched_game_name"] = alt.get("game_name", "")
                    m["confidence"] = alt.get("confidence", 0)
                    found = True
                    logger.info(f"[flex-import] 渠道过滤: '{m['candidate']}' {m['matched_game_id']}→{alt_gid}")
                    break

            if not found:
                errors.append({
                    "row": row_num,
                    "error": f"游戏名 '{m['candidate']}' 匹配到 {m['matched_game_id']}，但该游戏不在渠道 {channel_id} 的原始流水表中，且候选列表中也无匹配项",
                })
                rows[i]["game_id"] = None
                continue

        if m["status"] == "low" and matched_gid in valid_gids:
            errors.append({
                "row": row_num,
                "error": f"游戏名 '{m['candidate']}' 匹配置信度低 ({m['confidence']}%), 请手动确认",
            })

        rows[i]["game_id"] = matched_gid
        rows[i]["matched_game_name"] = m["matched_game_name"]
        rows[i]["match_confidence"] = m["confidence"]
        rows[i]["match_status"] = m["status"]
        rows[i]["match_candidates"] = m["candidates"]

    return match_results, errors


def _to_decimal(v) -> Decimal | None:
    if v is None:
        return None
    try:
        d = Decimal(str(v))
        return d if d >= 0 else None
    except Exception:
        return None


async def import_flexible_data(
    db: AsyncSession,
    rows: list[dict],
    channel_id: int,
    month: str,
    column_mapping: dict[str, str],
    selected_indices: set[int] | None = None,
) -> dict:
    from datetime import datetime as _dt
    from .lock_service import write_lock_inline

    deduction_map = {
        "vouchers": "vouchers", "test": "test",
        "welfare": "welfare", "bad_debt": "bad_debt",
    }

    now = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
    imported_deductions = 0
    imported_configs = 0
    imported_locks = 0
    skipped = 0
    skipped_settlements = 0

    # Guard: detect duplicate (channel_id, game_id, month) combos among selected rows
    seen = {}
    dupes = []
    for idx, record in enumerate(rows):
        if selected_indices is not None and idx not in selected_indices:
            continue
        gid = record.get("game_id")
        if not gid:
            continue
        rm = month or record.get("month", "")
        key = (channel_id, gid, rm)
        if key in seen:
            dupes.append((seen[key][0], idx, channel_id, gid, rm))
        else:
            seen[key] = (idx, record)

    if dupes:
        dup_details = "; ".join(
            f"行{first+1}与行{second+1}同时匹配渠道{ch} 游戏'{g}' 月份'{m}'"
            for first, second, ch, g, m in dupes
        )
        raise ValueError(f"导入数据中存在重复匹配的游戏条目，请修正后重新导入: {dup_details}")

    # ── Guard: reject import if any (channel, game, month) has an existing lock ──
    frozen_keys = []
    for idx, record in enumerate(rows):
        if selected_indices is not None and idx not in selected_indices:
            continue
        gid = record.get("game_id")
        if not gid:
            continue
        rm = month or record.get("month", "")
        if not rm:
            continue
        frozen_keys.append((channel_id, gid, rm))

    if frozen_keys:
        frozen_rows = (await db.execute(
            select(models.ChannelLock.channel_id, models.ChannelLock.game_id,
                   models.ChannelLock.month)
            .where(
                tuple_(models.ChannelLock.channel_id,
                       models.ChannelLock.game_id,
                       models.ChannelLock.month).in_(frozen_keys),
                or_(
                    models.ChannelLock.locked_real_revenue.isnot(None),
                    models.ChannelLock.locked_settlement_amount.isnot(None),
                ),
            )
        )).all()
        logger.info(f"[flex-import] 锁定守卫: 检查 {len(frozen_keys)} 键, 命中 {len(frozen_rows)} 锁")
        if frozen_rows:
            frozen_details = "; ".join(
                f"渠道{ch} 游戏'{g}' 月份{m}"
                for ch, g, m in frozen_rows
            )
            logger.warning(f"[flex-import] 锁定守卫拒绝导入: {frozen_details}")
            raise ValueError(
                f"以下数据已被锁定，无法通过弹性导入修改。请先解锁后再导入: {frozen_details}"
            )

    # ── Guard: only process games that exist in raw_settlements ──
    rs_keys = []
    for idx, record in enumerate(rows):
        if selected_indices is not None and idx not in selected_indices:
            continue
        gid = record.get("game_id")
        if not gid:
            continue
        rm = month or record.get("month", "")
        if not rm:
            continue
        rs_keys.append((channel_id, gid, rm))

    rs_map: dict[tuple, bool] = {}
    if rs_keys:
        rs_rows = (await db.execute(
            select(models.RawSettlement.channel_id, models.RawSettlement.game_id, models.RawSettlement.month)
            .where(
                tuple_(models.RawSettlement.channel_id, models.RawSettlement.game_id, models.RawSettlement.month).in_(rs_keys)
            )
        )).all()
        rs_map = {(r.channel_id, r.game_id, r.month): True for r in rs_rows}

    # ── Batch prefetch: Deductions + ChannelLocks ──
    import_keys = []
    for idx, record in enumerate(rows):
        if selected_indices is not None and idx not in selected_indices:
            continue
        gid = record.get("game_id")
        if not gid:
            continue
        rm = month or record.get("month", "")
        if not rm:
            continue
        import_keys.append((gid, rm))

    ded_map: dict[tuple, models.Deduction] = {}
    if import_keys:
        ded_rows = (await db.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == channel_id,
                tuple_(models.Deduction.game_id, models.Deduction.month).in_(import_keys),
            )
        )).scalars().all()
        ded_map = {(d.game_id, d.month): d for d in ded_rows}

    # Detect total_deductions fallback: if total_deductions is mapped but none of the
    # four individual deduction fields are, use it as the sole deduction source.
    mapped_fields = set(column_mapping.values()) if column_mapping else set()
    use_total_ded = (
        "total_deductions" in mapped_fields
        and not any(k in mapped_fields for k in ("vouchers", "test", "welfare", "bad_debt"))
    )

    # ── Process each row ──
    for idx, record in enumerate(rows):
        if selected_indices is not None and idx not in selected_indices:
            continue

        game_id = record.get("game_id")
        if not game_id:
            continue

        row_month = month or record.get("month", "")
        if not row_month:
            continue

        # ── Deductions (UPDATE only — 禁止新建) ──
        existing_ded = ded_map.get((game_id, row_month))
        if existing_ded:
            for src, dst in deduction_map.items():
                val = record.get(src)
                if val is not None:
                    setattr(existing_ded, dst, Decimal(str(val)))

            if use_total_ded:
                td = record.get("total_deductions")
                if td is not None:
                    existing_ded.vouchers = Decimal(str(td))

            imported_deductions += 1
        else:
            skipped += 1
            logger.info(f"[flex-import] Deduction skip: channel={channel_id} game={game_id} month={row_month} — 无已存在行")

        # ── Split config upsert ──
        split_rate = record.get("split_rate")
        channel_fee = record.get("channel_fee_rate")
        tax = record.get("tax_rate")
        if any(v is not None for v in (split_rate, channel_fee, tax)):
            await save_income_split_config(
                db, channel_id, game_id, row_month,
                split_rate=_to_decimal(split_rate) if split_rate is not None else None,
                channel_fee_rate=_to_decimal(channel_fee) if channel_fee is not None else None,
                tax_rate=_to_decimal(tax) if tax is not None else None,
            )
            imported_configs += 1

        # ── Lock: 写入 ChannelLock (real_revenue + settlement_amount) ──
        lock_fields = [("real_revenue", "real_revenue"), ("settlement_amount", "settlement_amount")]
        for src, lock_field in lock_fields:
            val = record.get(src)
            if val is not None:
                await write_lock_inline(
                    db, "channel", channel_id, game_id, row_month,
                    lock_field, Decimal(str(val)), now=now,
                )
                imported_locks += 1

    await db.commit()

    logger.info(
        f"[flex-import] 导入完成: deductions={imported_deductions} "
        f"configs={imported_configs} locks={imported_locks} skipped_ded={skipped}"
    )
    return {
        "imported_deductions": imported_deductions,
        "imported_transactions": 0,
        "imported_configs": imported_configs,
        "imported_locks": imported_locks,
        "skipped": skipped,
        "skipped_settlements": skipped_settlements,
    }


# ── Import comparison (moved from settlement_service) ──


async def compare_imported_rows(
    db: AsyncSession,
    rows: list[dict],
    channel_id: int,
    month: str,
    column_mapping: dict[str, str],
) -> list[dict]:
    """对比弹性导入行 vs 数据库已有数据。

    弹性导入只写 Deduction + IncomeSplitConfig，所以只对比：
      - vouchers / test / welfare / bad_debt ← Deduction
      - split_rate / channel_fee_rate / tax_rate ← IncomeSplitConfig
    raw_revenue/real_revenue/settlement_amount 来自原始流水表，弹性导入不写入，不作对比。
    """
    from decimal import Decimal as D

    compare_fields = [
        ("raw_revenue", "原始流水"), ("real_revenue", "真实流水"),
        ("settlement_amount", "结算金额"),
        ("vouchers", "代金券"), ("test", "测试"),
        ("welfare", "福利币"), ("bad_debt", "坏账"),
        ("split_rate", "分成比例"), ("channel_fee_rate", "通道费率"),
        ("tax_rate", "税率"),
    ]

    mapped_fields = {v for v in (column_mapping or {}).values() if v and v != "ignore"}
    field_defs = [(k, lbl) for k, lbl in compare_fields if k in mapped_fields]

    game_ids = [r.get("game_id") for r in rows if r.get("game_id")]
    if not game_ids:
        return []

    has_month = bool(month)
    ded_map: dict[tuple, models.Deduction] = {}
    lock_map: dict[tuple, models.ChannelLock] = {}
    month_closed_flag = False

    if has_month:
        ded_rows = (await db.execute(
            select(models.Deduction).where(
                models.Deduction.channel_id == channel_id,
                models.Deduction.game_id.in_(game_ids),
                models.Deduction.month == month,
            )
        )).scalars().all()
        ded_map = {(d.game_id, d.month): d for d in ded_rows}

        lock_rows = (await db.execute(
            select(models.ChannelLock).where(
                models.ChannelLock.channel_id == channel_id,
                models.ChannelLock.game_id.in_(game_ids),
                models.ChannelLock.month == month,
            )
        )).scalars().all()
        lock_map = {(lk.game_id, lk.month): lk for lk in lock_rows}

        month_closed_flag = (await db.execute(
            select(models.MonthlyClose.id).where(models.MonthlyClose.month == month)
        )).scalar_one_or_none() is not None

    # ── Batch: game discount_rate + IncomeSplitConfig (for current settlement values) ──
    game_rows = (await db.execute(
        select(models.Game.game_id, models.Game.discount_rate)
        .where(models.Game.game_id.in_(game_ids))
    )).all()
    discount_map = {g.game_id: float(g.discount_rate) for g in game_rows}

    cfg_map_sub: dict[tuple, models.IncomeSplitConfig] = {}
    if has_month:
        ms2, me2 = month_bounds(month)
        cfg_rows = (await db.execute(
            select(models.IncomeSplitConfig).where(
                models.IncomeSplitConfig.channel_id == channel_id,
                models.IncomeSplitConfig.game_id.in_(game_ids),
                models.IncomeSplitConfig.effective_from < me2,
                or_(models.IncomeSplitConfig.effective_to.is_(None),
                    models.IncomeSplitConfig.effective_to >= ms2),
            ).order_by(models.IncomeSplitConfig.effective_from.desc(), models.IncomeSplitConfig.id.desc())
        )).scalars().all()
        for c in cfg_rows:
            if c.game_id not in cfg_map_sub:
                cfg_map_sub[c.game_id] = c

    results = []
    for row in rows:
        gid = row.get("game_id")
        game_name = row.get("game_name") or row.get("matched_game_name", "")

        ded_row = ded_map.get((gid, month)) if has_month else None
        lock_row = lock_map.get((gid, month)) if has_month else None

        # Deduction old values (use is not None, not truthiness — 0 is valid)
        old_vouchers = float(ded_row.vouchers) if ded_row is not None and ded_row.vouchers is not None else None
        old_test = float(ded_row.test) if ded_row is not None and ded_row.test is not None else None
        old_welfare = float(ded_row.welfare) if ded_row is not None and ded_row.welfare is not None else None
        old_bad_debt = float(ded_row.bad_debt) if ded_row is not None and ded_row.bad_debt is not None else None

        # Config old values (use deduped cfg_map_sub, not per-row query)
        cfg_row = cfg_map_sub.get(gid) if has_month else None
        old_split = float(cfg_row.split_rate) if cfg_row else None
        old_cf = float(cfg_row.channel_fee_rate) if cfg_row else None
        old_tax = float(cfg_row.tax_rate) if cfg_row else None

        # 当前结算值（先算出来，用于 old_map 的 raw_revenue/real_revenue/settlement_amount）
        old_raw = old_real = old_settlement = None
        if has_month and gid:
            inp = await hydrate_formula_input(db, "channel", channel_id, gid, month)
            if inp:
                locked_r = lock_row.locked_real_revenue if lock_row else None
                locked_a = lock_row.locked_settlement_amount if lock_row else None
                old_raw = float(inp.raw_revenue)
                rv, sa = compute_settlement(
                    raw_revenue=inp.raw_revenue, discount_rate=inp.discount_rate,
                    total_deductions=inp.total_deductions,
                    split_rate=cfg_row.split_rate if cfg_row else D("0"),
                    channel_fee_rate=cfg_row.channel_fee_rate if cfg_row else D("0"),
                    tax_rate=cfg_row.tax_rate if cfg_row else D("0"),
                    locked_real_revenue=locked_r, locked_settlement_amount=locked_a,
                    direction="income",
                )
                old_real = float(rv)
                old_settlement = float(sa) if sa is not None else None

        old_map = {
            "raw_revenue": old_raw, "real_revenue": old_real,
            "settlement_amount": old_settlement,
            "vouchers": old_vouchers, "test": old_test,
            "welfare": old_welfare, "bad_debt": old_bad_debt,
            "split_rate": old_split, "channel_fee_rate": old_cf, "tax_rate": old_tax,
        }

        fields = {}
        for fkey, flabel in field_defs:
            new_val = row.get(fkey)
            old_val = old_map.get(fkey)

            new_val_f = None
            if new_val is not None:
                try:
                    new_val_f = float(new_val)
                except (ValueError, TypeError):
                    new_val_f = None

            changed = False
            if old_val is not None and new_val_f is not None:
                changed = abs(old_val - new_val_f) > 0.001
            elif old_val is not None or new_val_f is not None:
                changed = True

            fields[fkey] = {"label": flabel, "old": old_val, "new": new_val_f, "changed": changed}

        # ChannelLock info
        has_snapshot = lock_row is not None
        snapshot_info = None
        if has_snapshot:
            snapshot_info = {
                "lock_created_at": lock_row.created_at,
                "locked_real_revenue": float(lock_row.locked_real_revenue) if lock_row.locked_real_revenue is not None else None,
                "locked_settlement_amount": float(lock_row.locked_settlement_amount) if lock_row.locked_settlement_amount is not None else None,
            }

        # 当前结算值（复用上面已计算的值）
        settlement_current = None
        if old_raw is not None:
            settlement_current = {
                "raw_revenue": old_raw,
                "real_revenue": old_real,
                "settlement_amount": old_settlement,
            }

        results.append({
            "game_name": game_name,
            "game_id": gid,
            "fields": fields,
            "settlement_current": settlement_current,
            "has_snapshot": has_snapshot,
            "snapshot_info": snapshot_info,
        })

    # Detect duplicate game_ids
    gid_counts = {}
    for r in results:
        gid = r["game_id"]
        if gid:
            gid_counts[gid] = gid_counts.get(gid, 0) + 1
    for r in results:
        r["is_duplicate"] = r["game_id"] and gid_counts.get(r["game_id"], 0) > 1

    return results
