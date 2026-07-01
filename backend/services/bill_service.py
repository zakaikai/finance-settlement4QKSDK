"""Excel bill/invoice generation and CSV export for settlement data."""
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, Side, PatternFill,
)
from openpyxl.utils import get_column_letter


# ── Colours / styles ──

_TITLE_FONT = Font(name="微软雅黑", size=18, bold=True, color="1a1a2e")
_SUBTITLE_FONT = Font(name="微软雅黑", size=12, bold=True, color="333333")
_LABEL_FONT = Font(name="微软雅黑", size=11, color="444444")
_HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="ffffff")
_DATA_FONT = Font(name="微软雅黑", size=10, color="222222")
_SUMMARY_FONT = Font(name="微软雅黑", size=10, bold=True, color="222222")
_NOTE_FONT = Font(name="微软雅黑", size=9, color="888888")

_HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
_SUMMARY_FILL = PatternFill("solid", fgColor="f0f0f0")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

_NUM_FMT = '#,##0.00'
_PCT_FMT = '0.00%'


# ── Column definitions ──
# Identity columns per direction, followed by data columns from FIELD_DEFS.

_INCOME_ID_COLS = [
    ("序号",      "seq",          False, False),
    ("游戏编号",   "game_id",      False, False),
    ("游戏名称",   "game_name",    False, False),
    ("收入方名称", "channel_name", False, False),
    ("项目编号",   "project_code", False, False),
    ("项目名称",   "project_name", False, False),
]

_PAYMENT_ID_COLS = [
    ("序号",      "seq",            False, False),
    ("游戏编号",   "game_id",        False, False),
    ("游戏名称",   "game_name",      False, False),
    ("付款方名称", "publisher_name", False, False),
    ("项目编号",   "project_code",   False, False),
    ("项目名称",   "project_name",   False, False),
]

from .field_definitions import get_bill_columns as _get_bill_columns
_ID_KEYS = {"game_id", "game_name"}
_BILL_DATA_COLS = [c for c in _get_bill_columns() if c[1] not in _ID_KEYS]

_INCOME_COLUMNS = _INCOME_ID_COLS + _BILL_DATA_COLS
_PAYMENT_COLUMNS = _PAYMENT_ID_COLS + _BILL_DATA_COLS

_COLUMN_MAP = {
    "income": _INCOME_COLUMNS,
    "payment": _PAYMENT_COLUMNS,
}

_INCOME_WIDTHS = [6, 14, 16, 16, 14, 18] + [16, 16, 14, 12, 12, 12, 14, 12, 12, 10, 12, 16]
_PAYMENT_WIDTHS = [6, 14, 16, 16, 14, 18] + [16, 16, 14, 12, 12, 12, 14, 12, 12, 10, 12, 16]

_WIDTH_MAP = {
    "income": _INCOME_WIDTHS,
    "payment": _PAYMENT_WIDTHS,
}


def _style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _style_data_cell(cell, is_pct=False, is_money=False):
    cell.font = _DATA_FONT
    cell.border = _THIN_BORDER
    if is_money or is_pct:
        cell.alignment = _RIGHT
        cell.number_format = _PCT_FMT if is_pct else _NUM_FMT
    else:
        cell.alignment = _CENTER


def _write_party_info(ws, start_row, label, party):
    """Write party info block starting at start_row, return next available row."""
    rows = [
        (f"{label}：{party.name}", _SUBTITLE_FONT),
        (f"地  址：{party.address}", _LABEL_FONT),
        (f"电  话：{party.phone or ''}", _LABEL_FONT),
        (f"开户银行：{party.bank_name}", _LABEL_FONT),
        (f"银行账号：{party.bank_account}", _LABEL_FONT),
        (f"税  号：{party.tax_id}", _LABEL_FONT),
    ]
    for i, (text, font) in enumerate(rows):
        cell = ws.cell(row=start_row + i, column=1, value=text)
        cell.font = font
    return start_row + len(rows)


# ── Unified Excel bill generator ──

async def generate_settlement_bill(
    party_a,
    period: str,
    party_b,
    rows: list[dict],
    title: str = "收入结算对账单",
    mode: str = "income",
) -> BytesIO:
    """Build a formatted settlement bill Excel workbook.

    Parameters
    ----------
    party_a      — 甲方 PartyInfo ORM object
    period       : str — billing period display string
    party_b      — 乙方 PartyInfo ORM object
    rows         : list[dict] — settlement data rows
    title        : str — document title (e.g. "收入结算对账单")
    mode         : str — "income" or "payment" (determines columns)
    """
    cols = _COLUMN_MAP.get(mode, _INCOME_COLUMNS)
    widths = _WIDTH_MAP.get(mode, _INCOME_WIDTHS)
    col_count = len(cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"
    ws.sheet_properties.tabColor = "1a1a2e"

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    today = date.today()
    bill_no = f"ST-{today.strftime('%Y%m%d')}-{datetime.now().strftime('%H%M')}"

    # ── Title ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 42

    # ── Info block ──
    row = 3
    ws.cell(row=row, column=1, value=f"对账单编号:  {bill_no}").font = _LABEL_FONT
    row = 4
    ws.cell(row=row, column=1, value=f"对账周期:    {period}").font = _LABEL_FONT
    row = 5
    ws.cell(row=row, column=1, value=f"生成日期:    {today.strftime('%Y年%m月%d日')}").font = _LABEL_FONT

    # ── Parties ──
    row = 7
    row = _write_party_info(ws, row, "甲  方", party_a)
    row += 1
    row = _write_party_info(ws, row, "乙  方", party_b)

    # ── Data table headers ──
    header_row = row + 1
    for c, (header_name, *_) in enumerate(cols, 1):
        ws.cell(row=header_row, column=c, value=header_name)
    _style_header_row(ws, header_row, col_count)
    ws.row_dimensions[header_row].height = 28

    # ── Data rows ──
    for idx, r in enumerate(rows, 1):
        data_row = header_row + idx
        for c, (_, field, is_money, is_pct) in enumerate(cols, 1):
            if field == "seq":
                v = idx
            else:
                v = r.get(field)
            cell = ws.cell(row=data_row, column=c, value=v if v is not None else "")
            _style_data_cell(cell, is_pct=is_pct, is_money=is_money)
        ws.row_dimensions[data_row].height = 22

    total_data_rows = len(rows)
    data_end_row = header_row + total_data_rows

    # ── Summary row ──
    summary_row = data_end_row + 2
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
    sum_label = ws.cell(row=summary_row, column=1, value="合  计")
    sum_label.font = _SUMMARY_FONT
    sum_label.fill = _SUMMARY_FILL
    sum_label.alignment = _CENTER
    sum_label.border = _THIN_BORDER
    for c in range(2, 6):
        ws.cell(row=summary_row, column=c).fill = _SUMMARY_FILL
        ws.cell(row=summary_row, column=c).border = _THIN_BORDER

    sum_values = {}
    for r in rows:
        for _, field, is_money, _ in cols:
            if is_money:
                val = r.get(field)
                if val is not None:
                    sum_values[field] = sum_values.get(field, 0) + val

    for c, (_, field, is_money, _) in enumerate(cols, 1):
        if not is_money or field == "seq":
            continue
        val = round(sum_values.get(field, 0), 2)
        cell = ws.cell(row=summary_row, column=c, value=val)
        cell.font = _SUMMARY_FONT
        cell.fill = _SUMMARY_FILL
        cell.alignment = _RIGHT
        cell.border = _THIN_BORDER
        cell.number_format = _NUM_FMT

    # ── Notes ──
    note_row = summary_row + 2
    ws.cell(row=note_row, column=1, value="备注：").font = Font(name="微软雅黑", size=10, bold=True, color="666666")
    notes = [
        "1. 本对账单仅作为双方对账参考，不作为最终结算凭证。",
        "2. 如有异议，请于收到本对账单后 5 个工作日内联系我方。",
    ]
    for i, n in enumerate(notes, 1):
        ws.cell(row=note_row + i, column=1, value=n).font = _NOTE_FONT

    # ── Print settings ──
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Thin wrappers (backward compatibility) ──

async def generate_channel_bill(party_a, month, party_b, rows):
    return await generate_settlement_bill(
        party_a, month, party_b, rows,
        title="收 入 结 算 对 账 单", mode="income",
    )


async def generate_publisher_bill(party_a, month, party_b, rows):
    return await generate_settlement_bill(
        party_a, month, party_b, rows,
        title="付 款 结 算 对 账 单", mode="payment",
    )
