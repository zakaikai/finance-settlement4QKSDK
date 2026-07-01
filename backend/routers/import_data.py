from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook
import os
import json
import tempfile
import logging

logger = logging.getLogger(__name__)

from ..database import get_db
from .. import models
from ..services import template_defs as tdefs
from ..services import template_import as timp
from ..services import flexible_import as fimp
from ..services.settlement_service import compare_imported_rows as build_comparison
from ..utils.upload import saved_upload

# ── Synonym dictionary persistence ──
_SYNONYM_DICT_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "column_synonyms.json")

def _load_synonym_dict() -> dict:
    """Return merged dictionary: defaults as base, persisted file as overlay."""
    result = dict(tdefs._COLUMN_SYNONYMS)
    if os.path.exists(_SYNONYM_DICT_PATH):
        try:
            with open(_SYNONYM_DICT_PATH, "r", encoding="utf-8") as f:
                result.update(json.load(f))
        except Exception:
            pass
    return result

def _save_synonym_dict(data: dict) -> None:
    os.makedirs(os.path.dirname(_SYNONYM_DICT_PATH), exist_ok=True)
    with open(_SYNONYM_DICT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

router = APIRouter(prefix="/api/import", tags=["导入"])

TEMPLATES_DIR = os.environ.get(
    "TEMPLATES_DIR",
    os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
)


@router.get("/templates")
async def list_templates():
    return {
        "templates": [
            {"type": k, "label": v["label"], "columns": v["columns"]}
            for k, v in tdefs.TEMPLATE_DEFS.items()
        ]
    }


@router.get("/templates/{template_type}/download")
async def download_template(template_type: str):
    tdef = tdefs.TEMPLATE_DEFS.get(template_type)
    if not tdef:
        raise HTTPException(status_code=404, detail=f"模板 '{template_type}' 不存在")

    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    path = os.path.join(TEMPLATES_DIR, f"{template_type}.xlsx")

    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = tdef["label"]
        ws.append(tdef["columns"])
        wb.save(path)

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{template_type}.xlsx",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@router.post("/preview")
async def preview_import(
    template_type: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload and preview import data with validation + conflict detection."""
    tdef = tdefs.TEMPLATE_DEFS.get(template_type)
    if not tdef:
        raise HTTPException(status_code=400, detail=f"未知模板: {template_type}")

    async with saved_upload(file) as tmp_path:
        result = await timp.parse_excel(tmp_path, template_type)
        if "error" in result:
            raise HTTPException(status_code=400, detail=result["error"])

        rows = result["rows"]
        all_errors = list(result["errors"])

        # FK resolution
        fk_result = await timp.resolve_foreign_keys(db, template_type, rows)
        all_errors.extend(fk_result["errors"])

        # Value validation (only if no FK errors)
        if not fk_result["errors"]:
            val_result = await timp.validate_values(template_type, rows)
            all_errors.extend(val_result["errors"])

        # Conflict detection
        conflict_result = await timp.check_conflicts(db, template_type, rows)

        return {
            "template_type": template_type,
            "total_rows": result["total_rows"],
            "preview_rows": rows[:5],
            "errors": all_errors,
            "has_conflict": conflict_result["has_conflict"],
            "conflict_count": conflict_result["conflict_count"],
        }



@router.post("/confirm")
async def confirm_import(
    template_type: str = Form(...),
    file: UploadFile = File(...),
    overwrite: str = Form("false"),
    db: AsyncSession = Depends(get_db),
):
    """Confirm import: write validated data to database."""
    overwrite_flag = overwrite.lower() in ("true", "1")
    logger.info(f"[import-confirm] type={template_type} overwrite_flag={overwrite_flag}")
    tdef = tdefs.TEMPLATE_DEFS.get(template_type)
    if not tdef:
        raise HTTPException(status_code=400, detail=f"未知模板: {template_type}")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = await timp.parse_excel(tmp_path, template_type)
        if "error" in result:
            raise HTTPException(status_code=400, detail=result["error"])
        rows = result["rows"]
        if result["errors"]:
            raise HTTPException(status_code=400, detail={"errors": result["errors"]})

        fk_result = await timp.resolve_foreign_keys(db, template_type, rows)
        if fk_result["errors"]:
            raise HTTPException(status_code=400, detail={"errors": fk_result["errors"]})

        val_result = await timp.validate_values(template_type, rows)
        if val_result["errors"]:
            raise HTTPException(status_code=400, detail={"errors": val_result["errors"]})

        import_result = await timp.import_data(db, template_type, rows, overwrite_flag)
        return {"success": True, "imported": import_result["imported"]}
    finally:
        os.unlink(tmp_path)


# ── Flexible import (channel bill with column mapping) ──


@router.post("/flexible/preview")
async def flexible_preview(
    file: UploadFile = File(...),
    header_row: str = Form("1"),
):
    """Upload Excel → read header row → infer column mapping → preview sample data."""
    try:
        hdr_row = int(header_row)
    except ValueError:
        raise HTTPException(400, "header_row 必须为整数")

    async with saved_upload(file) as tmp_path:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        raw_rows = []

        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
            if i + 1 < hdr_row:
                continue
            if i + 1 == hdr_row:
                headers = [str(v).strip() if v is not None else "" for v in row]
                continue
            if all(v is None for v in row):
                continue
            raw_rows.append([str(v).strip() if v is not None else "" for v in row])
            if len(raw_rows) >= 10:
                break
        wb.close()

        if not headers:
            os.unlink(tmp_path)
            raise HTTPException(400, f"第 {hdr_row} 行无表头数据")

        mapping = tdefs.infer_column_mapping(headers)

        return {
            "headers": headers,
            "suggested_mapping": mapping,
            "preview_rows": raw_rows,
            "header_row": hdr_row,
        }



@router.post("/flexible/confirm")
async def flexible_confirm(
    file: UploadFile = File(...),
    header_row: str = Form("1"),
    month: str = Form(""),
    channel_id: str = Form("0"),
    column_mapping: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    """Parse Excel with mapping, fuzzy-match games, return for review. NO DB write."""
    try:
        hdr_row = int(header_row)
    except ValueError:
        raise HTTPException(400, "header_row 必须为整数")

    try:
        ch_id = int(channel_id)
    except ValueError:
        raise HTTPException(400, "channel_id 必须为整数")

    try:
        mapping = json.loads(column_mapping)
    except json.JSONDecodeError:
        raise HTTPException(400, "column_mapping JSON 格式错误")

    has_month_column = any(v == "month" for v in mapping.values())
    month_missing = not month and not has_month_column

    if ch_id <= 0:
        raise HTTPException(400, "channel_id 不能为空")

    async with saved_upload(file) as tmp_path:
        result = await fimp.parse_flexible_excel(tmp_path, mapping, hdr_row)
        rows = result["rows"]
        all_errors = list(result["errors"])

        game_name_col = None
        for ci_str, field_key in mapping.items():
            if field_key == "game_name":
                game_name_col = "game_name"
                break

        match_results = []
        if game_name_col:
            match_results, match_errors = await fimp.resolve_flexible_game_names(db, rows, game_name_col, ch_id, month)
            all_errors.extend(match_errors)

        # Build old-vs-new comparison
        comparison = await build_comparison(db, rows, ch_id, month, mapping)

        # Check if import month is ARAP-closed
        month_closed = False
        if month:
            month_closed = (await db.execute(
                select(models.MonthlyClose.id).where(models.MonthlyClose.month == month)
            )).scalar_one_or_none() is not None

        return {
            "rows": rows,
            "errors": all_errors,
            "total_rows": len(rows),
            "match_results": match_results,
            "comparison": comparison,
            "month_missing": month_missing,
            "month_closed": month_closed,
        }



@router.post("/flexible/import")
async def flexible_import(
    file: UploadFile = File(...),
    header_row: str = Form("1"),
    month: str = Form(""),
    channel_id: str = Form("0"),
    column_mapping: str = Form(...),
    selected_indices: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    """Parse + match + WRITE to DB (final import step).

    selected_indices: comma-separated zero-based row indices. Empty = import all.
    Auto-locks real_revenue and settlement_amount.
    Upserts split config if split_rate/channel_fee_rate/tax_rate provided.
    """
    try:
        hdr_row = int(header_row)
    except ValueError:
        raise HTTPException(400, "header_row 必须为整数")

    try:
        ch_id = int(channel_id)
    except ValueError:
        raise HTTPException(400, "channel_id 必须为整数")

    try:
        mapping = json.loads(column_mapping)
    except json.JSONDecodeError:
        raise HTTPException(400, "column_mapping JSON 格式错误")

    # Parse selected indices
    sel = None
    if selected_indices.strip():
        try:
            sel = {int(x.strip()) for x in selected_indices.split(",") if x.strip()}
        except ValueError:
            raise HTTPException(400, "selected_indices 格式错误")

    if ch_id <= 0:
        raise HTTPException(400, "channel_id 不能为空")

    ch = (await db.execute(
        select(models.ChannelCategory).where(models.ChannelCategory.channel_id == ch_id)
    )).scalar_one_or_none()
    if ch is None:
        raise HTTPException(400, f"渠道 ID {ch_id} 不存在")

    async with saved_upload(file) as tmp_path:
        result = await fimp.parse_flexible_excel(tmp_path, mapping, hdr_row)
        if not result["rows"]:
            err_detail = result["errors"][0] if result["errors"] else "未解析到任何数据行"
            raise HTTPException(400, detail=err_detail)

        rows = result["rows"]
        all_errors = list(result["errors"])

        game_name_col = None
        for ci_str, field_key in mapping.items():
            if field_key == "game_name":
                game_name_col = "game_name"
                break

        if game_name_col:
            match_results, match_errors = await fimp.resolve_flexible_game_names(db, rows, game_name_col, ch_id, month)
            all_errors.extend(match_errors)

        # Filter errors to only selected rows
        if sel is not None:
            all_errors = [e for e in all_errors if e.get("row", 0) - 2 in sel]

        # Allow low-confidence game matches through — user has manually reviewed
        blocking = [e for e in all_errors
                    if "置信度低" not in str(e.get("error", ""))]
        if blocking:
            raise HTTPException(400, detail={"errors": blocking})

        try:
            import_result = await fimp.import_flexible_data(db, rows, ch_id, month, mapping, sel)
        except ValueError as e:
            raise HTTPException(400, detail=str(e))

        return {
            "success": True,
            "imported_deductions": import_result["imported_deductions"],
            "imported_transactions": import_result["imported_transactions"],
            "imported_configs": import_result.get("imported_configs", 0),
            "total_rows": len(rows) if sel is None else len(sel),
        }



# ── Field definitions API (for frontend consumption) ──

@router.get("/flexible/field-definitions")
async def get_field_definitions():
    """Return canonical field definitions for frontend dropdowns etc."""
    from ..services.field_definitions import FIELD_DEFS
    return {
        "fields": [
            {"key": f["key"], "label": f["label"], "is_money": f["is_money"], "is_pct": f["is_pct"]}
            for f in FIELD_DEFS
        ]
    }


# ── Synonym dictionary management ──

@router.get("/flexible/dictionary")
async def export_synonym_dictionary():
    """Export the current column synonym dictionary as JSON."""
    data = _load_synonym_dict()
    return {"data": data, "fields": [f[0] for f in tdefs.FLEXIBLE_FIELD_DEFS]}


@router.post("/flexible/dictionary")
async def import_synonym_dictionary(
    file: UploadFile = File(...),
):
    """Upload a JSON file to replace the column synonym dictionary."""
    try:
        content = await file.read()
        data = json.loads(content.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        raise HTTPException(400, "文件格式错误，需要 UTF-8 编码的 JSON 文件")

    if not isinstance(data, dict):
        raise HTTPException(400, "JSON 格式错误，顶层必须为对象 {field_key: [synonyms, ...]}")

    valid_keys = {f[0] for f in tdefs.FLEXIBLE_FIELD_DEFS}
    invalid = [k for k in data if k not in valid_keys]
    if invalid:
        raise HTTPException(400, f"无效字段: {', '.join(invalid)}")

    for k, v in data.items():
        if not isinstance(v, list) or not all(isinstance(x, str) for x in v):
            raise HTTPException(400, f"字段 '{k}' 的值必须为字符串列表")

    _save_synonym_dict(data)
    tdefs._COLUMN_SYNONYMS.clear()
    tdefs._COLUMN_SYNONYMS.update(data)

    return {"success": True, "fields_updated": len(data)}
